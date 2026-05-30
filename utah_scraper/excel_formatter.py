import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys

def format_excel_report(leads, output_xlsx_path, output_pdf_path):
    """Styles the extracted leads list into a premium, Emerald-Green themed B2B leads pipeline and converts it to PDF."""
    print(f"\n[Excel Formatter] Generating styled report: {output_xlsx_path}...")
    
    # 1. Create DataFrame
    df = pd.DataFrame(leads)
    
    # Reorder columns for optimal readability
    column_order = [
        "company_name",
        "license_number",
        "original_issue_date",
        "qualifier_name",
        "registered_agent",
        "physical_address",
        "mailing_address",
        "classifications",
        "license_status"
    ]
    df = df.reindex(columns=column_order)
    
    # Rename columns to elegant display titles
    display_names = {
        "company_name": "Company Name",
        "license_number": "License Number",
        "original_issue_date": "Original Issue Date",
        "qualifier_name": "DOPL Qualifier",
        "registered_agent": "Registered Agent",
        "physical_address": "Physical Street Address",
        "mailing_address": "Mailing Address",
        "classifications": "Classifications",
        "license_status": "Status"
    }
    df = df.rename(columns=display_names)
    
    # 2. Save using openpyxl
    df.to_excel(output_xlsx_path, index=False)
    
    # 3. Apply Professional Styling
    wb = openpyxl.load_workbook(output_xlsx_path)
    ws = wb.active
    ws.title = "Utah Active Contractors"
    
    # Color Palette: Deep Emerald Theme
    header_fill = PatternFill(start_color="0F5132", end_color="0F5132", fill_type="solid") # Deep Forest Green
    zebra_fill = PatternFill(start_color="F4F9F4", end_color="F4F9F4", fill_type="solid")  # Ultra Light Mint Green
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="333333")
    bold_data_font = Font(name="Calibri", size=10, bold=True, color="111111")
    link_font = Font(name="Calibri", size=10, underline="single", color="0D6EFD") # Royal Blue Clickable
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D1E7DD") # Very light green/mint border
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Header Height and Alignment
    ws.row_dimensions[1].height = 35
    for col_idx in range(1, len(column_order) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = grid_border
        
    # Styles for data rows
    for row_idx in range(2, len(leads) + 2):
        ws.row_dimensions[row_idx].height = 32 # Premium spacious row height
        is_zebra = (row_idx % 2 == 1)
        current_fill = zebra_fill if is_zebra else white_fill
        
        for col_idx in range(1, len(column_order) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = current_fill
            cell.border = grid_border
            cell.font = data_font
            
            # Alignments & Formats based on column index
            col_letter = get_column_letter(col_idx)
            val = cell.value
            
            # Left align text columns
            if col_idx in [1, 4, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # Center align numbers, dates, statuses
            elif col_idx in [2, 3, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Style Company Name (Col 1) as bold
            if col_idx == 1:
                cell.font = bold_data_font
                
            # Make License Number (Col 2) a native clickable link pointing to search portal
            if col_idx == 2 and val:
                cell.value = f"{val} ↗"
                cell.hyperlink = "https://secure.utah.gov/llv/search/index.html"
                cell.font = link_font
                
            # Style Status (Col 9) in bold green
            if col_idx == 9:
                cell.font = Font(name="Calibri", size=10, bold=True, color="0F5132")
                
    # Auto-fit columns with safety padding
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            # Clean length calculation ignoring Excel hyperlink markup
            val_str = str(cell.value or '')
            if cell.hyperlink:
                val_str = val_str.replace(" ↗", "")
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # Save the polished workbook
    wb.save(output_xlsx_path)
    print("  -> Excel file saved and styled successfully.")
    
    # 4. Convert to landscape PDF using win32com (Excel COM Automation)
    print(f"[Excel Formatter] Exporting to landscape vector PDF: {output_pdf_path}...")
    try:
        import win32com.client
        
        # Absolute paths are required by Excel COM Automation
        abs_xlsx = os.path.abspath(output_xlsx_path)
        abs_pdf = os.path.abspath(output_pdf_path)
        
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.ScreenUpdating = False
        excel.DisplayAlerts = False
        
        try:
            wb_com = excel.Workbooks.Open(abs_xlsx)
            ws_com = wb_com.ActiveSheet
            
            # Set page setups for landscape fit-to-1-page wide
            ws_com.PageSetup.Orientation = 2 # 2 = xlLandscape
            ws_com.PageSetup.Zoom = False
            ws_com.PageSetup.FitToPagesWide = 1
            # Flow vertically over multiple pages naturally by not setting FitToPagesTall
            
            # Export
            ws_com.ExportAsFixedFormat(0, abs_pdf) # 0 = xlTypePDF
            wb_com.Close(False)
            print("  -> Landscape vector PDF generated successfully.")
        except Exception as com_err:
            print(f"  -> COM execution error during PDF export: {com_err}")
        finally:
            excel.Quit()
    except Exception as e:
        print(f"  -> win32com not available or PDF conversion failed: {e}")
        print("     (Verify that Excel is installed on your Windows machine to run local PDF conversion).")
