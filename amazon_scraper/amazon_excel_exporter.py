import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
import os
import re

class AmazonExcelExporter:
    """
    Exports scraped Amazon data into an executive-ready "Midnight Gold" workbook.
    Implements advanced copywriting title cleaning, increased row padding,
    left cell indentation, and a luxury visual design.
    """

    def __init__(self):
        pass

    def clean_title(self, title_str):
        """
        Cleans and shortens long, cluttered e-commerce SEO titles to make them
        look extremely clean, readable, and professional in Excel.
        E.g.: "ASUS ROG Strix G16 Gaming Laptop: 16” 165Hz FHD+, GeForce RTX 4060..."
        becomes: "ASUS ROG Strix G16"
        """
        if not title_str:
            return ""
        
        # 1. Clean curly quotes and weird characters to prevent Windows encoding bugs
        title_str = title_str.replace('”', '"').replace('“', '"').replace('’', "'").replace('‘', "'")
        title_str = title_str.replace('\xa0', ' ')
        title_str = re.sub(r'[^\x00-\x7F]+', ' ', title_str) # Strip any remaining non-ASCII glyphs
        
        # 2. Split by common delimiters used in SEO titles, now including colons (:)!
        parts = re.split(r'[:|,\-\(\[/\\]', title_str)
        
        # 3. Extract the first main part
        main_title = parts[0].strip()
        
        # 4. If the first part is too short (e.g. less than 12 chars), let's append the second part
        if len(main_title) < 12 and len(parts) > 1:
            main_title = f"{main_title} {parts[1].strip()}"
            
        # 5. Clean up redundant e-commerce marketing buzzwords
        cleaned_temp = main_title
        buzzwords = [
            r'\bGaming Laptop\b', r'\bGaming Notebook\b', r'\bLaptop\b', r'\bNotebook\b',
            r'\bComputer\b', r'\bSmart Phone\b', r'\bPhone\b', r'\bUnlocked\b',
            r'\bTitanium\b', r'\bBlack\b', r'\bSierra Blue\b', r'\bRenewed\b', r'\bPremium\b'
        ]
        for buzzword in buzzwords:
            cleaned_temp = re.sub(buzzword, '', cleaned_temp, flags=re.IGNORECASE).strip()
            
        # Safeguard: only apply buzzword cleaning if it doesn't leave the name too short/generic (e.g. Dell)
        if len(cleaned_temp) >= 8:
            main_title = cleaned_temp
            
        # 6. Clean up double spaces
        main_title = re.sub(r'\s+', ' ', main_title).strip()
        
        # 7. Truncate to maximum 40 characters so it fits neatly on a single line
        if len(main_title) > 40:
            main_title = main_title[:37] + "..."
            
        return main_title


    def export_data(self, data, output_path):
        """
        Transforms a list of phone/laptop dictionaries into a beautiful luxury-themed Excel workbook.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Amazon Tech Products"
        ws.views.sheetView[0].showGridLines = True
        
        # Columns (Neat headers using Link representations)
        headers = ["Product Title", "Price (USD)", "Status", "Product Link", "Image Link"]
        ws.append(headers)
        
        # Write Data
        for item in data:
            price = item.get("Price", 0.0)
            status = "Available" if price > 0 else "Out of Stock"
            
            # Clean and shorten the title before writing it to Excel
            raw_title = item.get("Title", "")
            cleaned_title = self.clean_title(raw_title)
            
            product_url = item.get("Product URL", "")
            image_url = item.get("Image URL", "")
            
            row_data = [
                cleaned_title,
                price,
                status,
                "View Product ↗" if product_url else "N/A",
                "View Image 📷" if image_url else "N/A"
            ]
            ws.append(row_data)
            
            # Set native Excel hyperlink properties (completely functional in all language configurations!)
            current_row = ws.max_row
            if product_url:
                ws.cell(row=current_row, column=4).hyperlink = product_url
            if image_url:
                ws.cell(row=current_row, column=5).hyperlink = image_url

        # Luxury "Midnight Gold" Palette Styling
        header_fill = PatternFill(start_color="1A202C", end_color="1A202C", fill_type="solid") # Midnight Dark Slate
        header_font = Font(name="Segoe UI", size=11, bold=True, color="D4AF37") # Premium Gold Color
        
        zebra_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid") # Off-White
        normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border_side = Side(border_style="thin", color="E2E8F0")
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Professional Left padding/indent to prevent text from touching cell lines
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=False, indent=1)
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        num_rows = len(data)
        data_start_row = 2
        data_end_row = num_rows + 1

        # Header Styling
        ws.row_dimensions[1].height = 30
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            # Align center for Price, Status, Links; Left align for Title
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                cell.alignment = align_center
            cell.border = cell_border

        # Data Row Styling (Airy padding, height 35)
        for r_idx in range(data_start_row, data_end_row + 1):
            ws.row_dimensions[r_idx].height = 35 # High-end airy row height
            row_fill = zebra_fill if r_idx % 2 == 0 else normal_fill
            
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.fill = row_fill
                cell.border = cell_border
                
                if c_idx == 1: # Title
                    cell.font = Font(name="Segoe UI", size=10)
                    cell.alignment = align_left
                elif c_idx == 2: # Price
                    cell.font = Font(name="Segoe UI", size=10)
                    cell.alignment = align_right
                    cell.number_format = '$#,##0.00'
                elif c_idx == 3: # Status
                    cell.font = Font(name="Segoe UI", size=10)
                    cell.alignment = align_center
                else: # Links (Columns 4 and 5)
                    cell.alignment = align_center
                    if cell.value == "N/A":
                        cell.font = Font(name="Segoe UI", size=10, italic=True, color="A0AEC0")
                    else:
                        cell.font = Font(name="Segoe UI", size=10, underline="single", color="3182CE")

        # Summary Row (Formulas)
        summary_row = data_end_row + 1
        ws.row_dimensions[summary_row].height = 30
        
        summary_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
        summary_font = Font(name="Segoe UI", size=10, bold=True, color="1A202C")
        double_bottom_border = Border(
            top=Side(border_style="thin", color="A0AEC0"),
            bottom=Side(border_style="double", color="D4AF37"), # Double Gold line
            left=thin_border_side,
            right=thin_border_side
        )

        ws.cell(row=summary_row, column=1, value="Average Price (of listed stock)").alignment = align_left
        ws.cell(row=summary_row, column=2, value=f'=SUMIF(B2:B{data_end_row}, ">0") / COUNTIF(B2:B{data_end_row}, ">0")').number_format = '$#,##0.00'
        ws.cell(row=summary_row, column=3, value="Total Models").alignment = align_center
        ws.cell(row=summary_row, column=4, value=f"=COUNTA(A{data_start_row}:A{data_end_row})").number_format = '0'
        ws.cell(row=summary_row, column=5, value="").alignment = align_center

        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=summary_row, column=c_idx)
            cell.fill = summary_fill
            cell.font = summary_font
            cell.border = double_bottom_border
            if c_idx == 2:
                cell.alignment = align_right
            elif c_idx in [3, 4, 5]:
                cell.alignment = align_center

        # Conditional Formatting for Out of Stock
        yellow_fill = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")
        yellow_font = Font(color="744210", bold=True)
        green_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
        green_font = Font(color="22543D", bold=True)

        ws.conditional_formatting.add(
            f"C{data_start_row}:C{data_end_row}",
            CellIsRule(operator="equal", formula=['"Available"'], fill=green_fill, font=green_font)
        )
        ws.conditional_formatting.add(
            f"C{data_start_row}:C{data_end_row}",
            CellIsRule(operator="equal", formula=['"Out of Stock"'], fill=yellow_fill, font=yellow_font)
        )

        # Embedded Price Chart (Top 10 priced items)
        chart_items = 0
        for item in data:
            if item.get("Price", 0.0) > 0:
                chart_items += 1
        
        chart_limit = min(chart_items, 10)
        if chart_limit > 1:
            chart = BarChart()
            chart.type = "col"
            chart.style = 11
            chart.title = "Amazon Product Price Comparison"
            chart.y_axis.title = "Price (USD)"
            chart.x_axis.title = "Product Title"
            
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=chart_limit + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=chart_limit + 1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.legend = None
            
            chart.width = 18
            chart.height = 10
            ws.add_chart(chart, "G2")

        # Auto-Fit Dimensions (Calculated based on actual display values, keeping columns compact)
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith('='):
                    val_str = "Average Price (of listed stock)"
                max_len = max(max_len, len(val_str))
                
            if col[0].column == 1:
                # Give product titles breathing room up to 45 chars
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 18), 45)
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        wb.save(output_path)
        print(f"Midnight Gold Excel report saved successfully to: {output_path}")

# Self-test
if __name__ == "__main__":
    exporter = AmazonExcelExporter()
    test_data = [
        {"Title": "ASUS ROG Strix G16 Gaming Laptop, 16” 165Hz FHD+, GeForce RTX 4060, Intel Core i7-13650HX, 16GB DDR5, 512GB PCIe SSD", "Price": 1299.00, "Product URL": "http://example.com", "Image URL": "http://example.com/img"},
        {"Title": "Apple iPhone 16 Pro Max, 256GB, Black Titanium - Unlocked", "Price": 1099.00, "Product URL": "http://example.com", "Image URL": "http://example.com/img"}
    ]
    exporter.export_data(test_data, "test_luxury_report.xlsx")
    if os.path.exists("test_luxury_report.xlsx"):
        os.remove("test_luxury_report.xlsx")
        print("Self-test clean and completed!")
