import sys
import os
import re
import time
import random
import argparse
import base64
import urllib.parse
from curl_cffi import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Force UTF-8 encoding on Windows console to avoid encoding crashes
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except AttributeError:
    pass

class B2BAgencyOutreachScraper:
    def __init__(self):
        # Impersonate Chrome to bypass basic firewalls and WAF blocks
        self.session = requests.Session(impersonate="chrome110")
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9,it;q=0.8,ro;q=0.7",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Referer": "https://www.google.com/"
        }
        # Ignore giant aggregator directories, social networks, or irrelevant platforms
        self.ignore_domains = [
            "facebook.com", "instagram.com", "linkedin.com", "youtube.com", "twitter.com", "pinterest.com", 
            "wikipedia.org", "google.com", "duckduckgo.com", "bing.com", "yahoo.com", "yandex.com",
            "clutch.co", "upwork.com", "fiverr.com", "indeed.com", "glassdoor.com", "g2.com", "sortlist.com",
            "paginegialle.it", "paginiaurii.ro", "yelp.com", "yellowpages.com", "tripadvisor.com",
            "github.com", "wix.com", "squarespace.com", "wordpress.com", "shopify.com", "behance.net", "dribbble.com"
        ]

    def is_ignored(self, domain):
        domain_low = domain.lower()
        return any(x in domain_low for x in self.ignore_domains)

    def clean_domain(self, url):
        try:
            parsed = urllib.parse.urlparse(url)
            domain = parsed.netloc.replace("www.", "")
            return domain
        except:
            return ""

    def search_duckduckgo(self, query, target_count):
        """Searches DuckDuckGo HTML interface for agency sites."""
        print(f"  [Search] Căutare pe DuckDuckGo pentru: {query}")
        links = []
        try:
            encoded_query = urllib.parse.quote_plus(query)
            url = f"https://html.duckduckgo.com/html/?q={encoded_query}"
            r = self.session.get(url, headers=self.headers, timeout=12)
            if r.status_code != 200:
                print(f"    [DuckDuckGo] Status code eroare: {r.status_code}")
                return []
                
            soup = BeautifulSoup(r.text, 'html.parser')
            for a in soup.select('a.result__url'):
                href = a.get('href', '')
                if 'uddg=' in href:
                    actual_url = href.split('uddg=')[1].split('&')[0]
                    actual_url = urllib.parse.unquote(actual_url)
                    if actual_url.startswith('http'):
                        links.append(actual_url)
            print(f"    [DuckDuckGo] Găsite {len(links)} link-uri brute.")
        except Exception as e:
            print(f"    [DuckDuckGo] Eroare la căutare: {e}")
        return links

    def search_bing(self, query, target_count):
        """Searches Bing for agency sites."""
        print(f"  [Search] Căutare pe Bing pentru: {query}")
        links = []
        try:
            encoded_query = urllib.parse.quote_plus(query)
            url = f"https://www.bing.com/search?q={encoded_query}"
            r = self.session.get(url, headers=self.headers, timeout=12)
            if r.status_code != 200:
                print(f"    [Bing] Status code eroare: {r.status_code}")
                return []
                
            soup = BeautifulSoup(r.text, 'html.parser')
            for a in soup.select('li.b_algo h2 a'):
                href = a.get('href', '')
                if 'bing.com/ck/a' in href:
                    parsed = urllib.parse.urlparse(href)
                    params = urllib.parse.parse_qs(parsed.query)
                    u_val = params.get('u', [''])[0]
                    if u_val:
                        b64_str = u_val[2:]
                        # Add base64 padding if needed
                        b64_str += '=' * (-len(b64_str) % 4)
                        try:
                            dec = base64.b64decode(b64_str).decode('utf-8', errors='ignore')
                            if dec.startswith('http'):
                                links.append(dec)
                        except:
                            pass
                elif href.startswith('http'):
                    links.append(href)
            print(f"    [Bing] Găsite {len(links)} link-uri brute.")
        except Exception as e:
            print(f"    [Bing] Eroare la căutare: {e}")
        return links

    def clean_leads_candidates(self, links, city, country):
        """Filters out major aggregators and formats candidates."""
        candidates = []
        seen_domains = set()
        for link in links:
            domain = self.clean_domain(link)
            if not domain or self.is_ignored(domain):
                continue
            if domain not in seen_domains:
                seen_domains.add(domain)
                # Formulate a clean company name from the domain name
                name_parts = domain.split(".")[0].replace("-", " ").replace("_", " ").title()
                candidates.append({
                    "name": name_parts,
                    "website": f"https://{domain}",
                    "location": f"{city}, {country}"
                })
        return candidates

    def clean_emails(self, emails, domain):
        """Filters out non-working or asset emails."""
        valid_emails = []
        for email in emails:
            email_low = email.strip().lower()
            # Regex check
            if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email_low):
                continue
            # Filter asset extensions
            if any(email_low.endswith(ext) for ext in [".png", ".jpg", ".jpeg", ".gif", ".svg", ".css", ".js"]):
                continue
            # Filter common false positives
            if any(x in email_low for x in ["bootstrap", "jquery", "w3.org", "example", "domain.com", "yourdomain"]):
                continue
            
            valid_emails.append(email_low)
            
        # Deduplicate and prioritize domain emails
        unique_emails = list(set(valid_emails))
        
        # Sort so that emails containing the domain or direct contact handles come first
        def email_priority(e):
            if domain in e:
                # Corporate domain email is high priority
                if any(handle in e for handle in ["info@", "contact@", "office@", "hello@", "team@", "agency@"]):
                    return 0
                return 1
            if any(handle in e for handle in ["info@", "contact@", "office@", "hello@", "team@"]):
                return 2
            return 3
            
        unique_emails.sort(key=email_priority)
        return unique_emails

    def extract_phones(self, text):
        """Extracts and validates phone numbers."""
        # Match various international and local phone formats
        phone_patterns = [
            r'\+?\d{1,4}[-.\s]?\(?\d{1,3}\)?[-.\s]?\d{3}[-.\s]?\d{4}', # standard US/UK format
            r'\+?\d{2,3}[-.\s]?\d{2,4}[-.\s]?\d{3,4}[-.\s]?\d{3,4}',  # Italian/Romanian format
            r'\+?\d{9,15}' # simple long number
        ]
        
        candidates = []
        for pat in phone_patterns:
            matches = re.findall(pat, text)
            for m in matches:
                # Strip spaces, hyphens, and dots to count digits
                digits = re.sub(r'\D', '', m)
                if 9 <= len(digits) <= 15:
                    candidates.append(m.strip())
                    
        return list(set(candidates))

    def extract_decidents(self, soup_text):
        """Scans website content for decision makers (CEO, Founder, Owner, Managing Director)."""
        decident_patterns = [
            r'(?:fondatore|amministratore|titolare|ceo|founder|director|fondator|owner|managing director)\s*[:\-]\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2})',
            r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2})\s*,\s*(?:fondatore|amministratore|titolare|ceo|founder|director|fondator|owner|managing director)',
            r'(?:fondatore|amministratore|titolare|ceo|founder|director|fondator|owner|managing\s+director)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2})'
        ]
        
        found_names = []
        for pat in decident_patterns:
            matches = re.finditer(pat, soup_text, re.IGNORECASE)
            for match in matches:
                name = match.group(1).strip()
                # Exclude common corporate titles or layout words caught by mistake
                name_low = name.lower()
                if any(x in name_low for x in ["contact", "about", "privacy", "policy", "terms", "agency", "marketing", "website", "company", "team", "career"]):
                    continue
                # Split and verify capitalization of name
                parts = name.split()
                if len(parts) >= 2:
                    found_names.append(name)
                    
        return list(set(found_names))

    def find_contact_subpages(self, homepage_url, soup):
        """Discovers relative or absolute links to Contact, About, or Team pages."""
        subpages = []
        keywords = ["contact", "about", "team", "staff", "chipa", "contatt", "despre", "noi", "privacy", "legal"]
        
        for a in soup.find_all('a', href=True):
            href = a['href'].strip()
            text = a.get_text().strip().lower()
            href_low = href.lower()
            
            is_match = False
            # Check if href or link text matches contact/about/team keywords
            if any(kw in href_low for kw in keywords):
                is_match = True
            elif any(kw in text for kw in keywords):
                is_match = True
                
            if is_match:
                # Exclude social media shares, external mailtos, or tel links
                if href_low.startswith('mailto:') or href_low.startswith('tel:') or href_low.startswith('javascript:'):
                    continue
                if any(x in href_low for x in ["facebook.com", "linkedin.com", "twitter.com", "instagram.com"]):
                    continue
                    
                absolute_url = urllib.parse.urljoin(homepage_url, href)
                # Keep it within the same domain
                if self.clean_domain(absolute_url) == self.clean_domain(homepage_url):
                    subpages.append(absolute_url)
                    
        return list(set(subpages))

    def crawl_agency_website(self, company_name, url):
        """Crawls home page and contact subpages of a website to extract real contact info."""
        print(f"    [Crawler] Scanăm site-ul: {url}...")
        domain = self.clean_domain(url)
        emails = []
        phones = []
        decidents = []
        linkedin_page = "N/A"
        source_url = url
        
        # 1. Fetch homepage
        try:
            r = self.session.get(url, headers=self.headers, timeout=12)
            if r.status_code not in [200, 301, 302]:
                # Retry with HTTP if HTTPS fails
                if url.startswith('https://'):
                    url_http = url.replace('https://', 'http://')
                    r = self.session.get(url_http, headers=self.headers, timeout=10)
                    
            html = r.text
            soup = BeautifulSoup(html, 'html.parser')
            text_content = soup.get_text()
            
            # Extract emails, phones, and decidents from homepage
            emails.extend(re.findall(r'[\w\.-]+@[\w\.-]+\.\w+', html))
            phones.extend(self.extract_phones(text_content))
            decidents.extend(self.extract_decidents(text_content))
            
            # Find LinkedIn page
            for a in soup.find_all('a', href=True):
                href = a['href']
                if "linkedin.com/company/" in href or "linkedin.com/in/" in href:
                    linkedin_page = href
                    
            # Find subpages to deep-scan (Contact, About Us, Team)
            subpages = self.find_contact_subpages(url, soup)
            print(f"      [Crawler] Descoperite {len(subpages)} pagini secundare (Contact/Team/About).")
            
            # Crawl up to 3 subpages
            for sub_url in subpages[:3]:
                try:
                    time.sleep(random.uniform(0.5, 1.2)) # Politeness delay
                    sub_r = self.session.get(sub_url, headers=self.headers, timeout=8)
                    sub_soup = BeautifulSoup(sub_r.text, 'html.parser')
                    sub_text = sub_soup.get_text()
                    
                    sub_emails = re.findall(r'[\w\.-]+@[\w\.-]+\.\w+', sub_r.text)
                    sub_phones = self.extract_phones(sub_text)
                    sub_decidents = self.extract_decidents(sub_text)
                    
                    if sub_emails:
                        emails.extend(sub_emails)
                        source_url = sub_url # Set source to the page we actually found contact info on
                    if sub_phones:
                        phones.extend(sub_phones)
                    if sub_decidents:
                        decidents.extend(sub_decidents)
                        
                    for a in sub_soup.find_all('a', href=True):
                        href = a['href']
                        if ("linkedin.com/company/" in href or "linkedin.com/in/" in href) and linkedin_page == "N/A":
                            linkedin_page = href
                except Exception as sub_err:
                    pass
                    
        except Exception as err:
            print(f"      [Crawler] Eșuat accesarea site-ului principal: {err}")
            
        # 2. Clean and validate extracted data
        cleaned_emails = self.clean_emails(emails, domain)
        final_email = cleaned_emails[0] if cleaned_emails else "N/A"
        
        # Format telephone
        final_phone = "N/A"
        if phones:
            # Clean and take the first phone number
            raw_phone = phones[0]
            clean_digits = re.sub(r'[^\d+]', '', raw_phone)
            # Reformat nice representation
            if clean_digits.startswith('0') and len(clean_digits) == 10:
                final_phone = f"{clean_digits[:4]} {clean_digits[4:7]} {clean_digits[7:]}"
            else:
                final_phone = raw_phone
                
        final_decident = decidents[0] if decidents else "N/A"
        
        return final_email, final_phone, final_decident, linkedin_page, source_url

    def run_pipeline(self, target_count):
        print("\n=======================================================")
        print("🚀 INIȚIALIZARE CĂUTARE GLOBALĂ B2B LEAD GENERATION")
        print(f"   Target total: {target_count} agenții din hub-uri globale")
        print("=======================================================\n")
        
        targets = [
            {"city": "New York", "country": "USA", "query": '"web design agency" OR "digital marketing agency" New York'},
            {"city": "London", "country": "UK", "query": '"web design agency" OR "digital marketing agency" London'},
            {"city": "Milano", "country": "Italy", "query": '"agenzia web" OR "agenzia marketing" Milano'},
            {"city": "București", "country": "Romania", "query": '"agentie web" OR "agentie marketing" Bucuresti'}
        ]
        
        leads = []
        leads_needed_per_city = (target_count // len(targets)) + 1
        
        for tgt in targets:
            city = tgt["city"]
            country = tgt["country"]
            query = tgt["query"]
            
            print(f"\n[*] Începem colectarea pentru: {city}, {country}...")
            
            # Attempt DuckDuckGo search first
            links = self.search_duckduckgo(query, leads_needed_per_city)
            # If DuckDuckGo fails or returns few results, fallback to Bing
            if len(links) < leads_needed_per_city:
                print(f"    [Fallback] DuckDuckGo a returnat puține rezultate. Încercăm Bing...")
                bing_links = self.search_bing(query, leads_needed_per_city)
                links = list(set(links + bing_links))
                
            candidates = self.clean_leads_candidates(links, city, country)
            print(f"    [Search] Identificate {len(candidates)} site-uri potențiale pentru {city}.")
            
            city_leads_count = 0
            for cand in candidates:
                if city_leads_count >= leads_needed_per_city or len(leads) >= target_count:
                    break
                    
                name = cand["name"]
                web = cand["website"]
                loc = cand["location"]
                
                print(f"\n  -> Analizăm [{len(leads) + 1}/{target_count}]: {name}")
                email, phone, decident, linkedin, source = self.crawl_agency_website(name, web)
                
                # We require at least an email or phone for a B2B prospect to be valid
                if email == "N/A" and phone == "N/A":
                    print(f"     [Ignorat] Nicio informație de contact reală găsită pe site. Sărim peste.")
                    continue
                    
                leads.append({
                    "company_name": name,
                    "location": loc,
                    "website": web,
                    "email": email,
                    "phone": phone,
                    "decident": decident,
                    "linkedin": linkedin,
                    "source_url": source
                })
                city_leads_count += 1
                print(f"     [CALIFICAT] {name} | Email: {email} | Tel: {phone} | Decident: {decident}")
                time.sleep(random.uniform(1.0, 2.5)) # Politeness delay between sites
                
            if len(leads) >= target_count:
                break
                
        return leads[:target_count]

    def format_excel_report(self, leads, output_path):
        """Formats the output into a premium HSL-styled Emerald Mint spreadsheet."""
        print(f"\n[Excel] Generăm registrul premium: {output_path}...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Outreach Agency Leads"
        
        # Gridlines must be visible
        ws.views.sheetView[0].showGridLines = True
        
        headers = [
            "Firma / Compania", "Locație / Oraș", "Site Web", 
            "E-mail Contact", "Telefon Contact", "Decident (CEO/Owner)", 
            "LinkedIn Companie", "Pagina Sursă Contact"
        ]
        
        # HSL Theme: Emerald Mint
        header_fill = PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid") # Dark Emerald
        zebra_fill = PatternFill(start_color="F2F7F5", end_color="F2F7F5", fill_type="solid")  # Mint Ice
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10, color="2D3748")
        bold_font = Font(name="Segoe UI", size=10, bold=True, color="1A202C")
        link_font = Font(name="Segoe UI", size=10, underline="single", color="1B4D3E") # Clickable Emerald Green
        
        thin_border_side = Side(border_style="thin", color="E2E8F0")
        grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Header Row Styling (Height = 35pt)
        ws.row_dimensions[1].height = 35
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = grid_border
            
        # Data Rows Styling (Height = 32pt - Spacious)
        for row_idx, lead in enumerate(leads, start=2):
            ws.row_dimensions[row_idx].height = 32
            row_fill = zebra_fill if (row_idx % 2 == 0) else white_fill
            
            row_data = [
                lead["company_name"], lead["location"], lead["website"],
                lead["email"], lead["phone"], lead["decident"],
                lead["linkedin"], lead["source_url"]
            ]
            
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.border = grid_border
                cell.font = data_font
                
                # Alignments
                if col_idx in [1, 2, 4, 6]: # Text/Text/Email/Decident
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else: # Website, Phone, LinkedIn, Source
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                # Values & Formatting
                if col_idx == 1:
                    cell.value = val
                    cell.font = bold_font
                elif col_idx == 3:
                    # Short interactive website link
                    cell.value = "Vizitează Site ↗"
                    cell.hyperlink = val
                    cell.font = link_font
                elif col_idx == 7:
                    # Short interactive LinkedIn link
                    if val != "N/A":
                        cell.value = "LinkedIn Profile ↗"
                        cell.hyperlink = val
                        cell.font = link_font
                    else:
                        cell.value = "N/A"
                elif col_idx == 8:
                    # Short interactive Source page link
                    if val.startswith('http'):
                        cell.value = "Pagina Sursă ↗"
                        cell.hyperlink = val
                        cell.font = link_font
                    else:
                        cell.value = val
                else:
                    cell.value = val
                    
        # Dynamic Auto-fitting columns
        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if cell.hyperlink:
                    val_str = val_str # Use the label text length
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
        wb.save(output_path)
        print(f"  -> Raport finalizat salvat în: {output_path}")

def main():
    parser = argparse.ArgumentParser(description="Global B2B Agency Lead Scraper")
    parser.add_argument("--count", type=int, default=30, help="Number of leads to scrape")
    args = parser.parse_args()
    
    scraper = B2BAgencyOutreachScraper()
    leads = scraper.run_pipeline(args.count)
    
    if not leads:
        print("\n[!] Nu s-au putut extrage lead-uri valide. Verificați conexiunea la internet sau interogările.")
        return
        
    # File locations
    filename = "PROSPECTE_AGENTII_OUTREACH.xlsx"
    local_dir = os.path.dirname(os.path.abspath(__file__))
    local_path = os.path.join(local_dir, filename)
    
    # Desktop path
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop", "Outreach_B2B")
    os.makedirs(desktop_dir, exist_ok=True)
    desktop_path = os.path.join(desktop_dir, filename)
    
    # Save spreadsheet local and copy to desktop
    scraper.format_excel_report(leads, local_path)
    
    try:
        import shutil
        shutil.copy(local_path, desktop_path)
        print(f"\n🚀 RAPORT FINALIZAT SALVAT PE DESKTOP: {desktop_path}")
        print(f"   Puteți deschide fișierul din folderul de pe Desktop: Outreach_B2B/{filename}")
    except Exception as e:
        print(f"   Eroare la salvarea pe Desktop: {e}")
        
    print("\n" + "=" * 55)
    print(f"🎉 SUCCES! Extrase {len(leads)} agenții B2B cu date reale.")
    print("=" * 55 + "\n")

if __name__ == "__main__":
    main()
