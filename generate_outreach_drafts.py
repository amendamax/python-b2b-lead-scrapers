import os
import re
import urllib.parse
from email.message import EmailMessage
from email.policy import SMTP
import openpyxl

def clean_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "", name).replace(" ", "_")

def get_email_templates(company_name, decident, language):
    # Determine salutation
    salutation = decident.split()[0] if (decident and decident != "N/A") else ""
    
    if language == "italian":
        salute = f"Ciao {salutation}" if salutation else "Gentile Team de"
        subject = f"Collaborazione Web Scraping & Automatizzazione Dati per {company_name}"
        body = f"""{salute} {company_name},

Vi contatto perché apprezzo molto i progetti digitali che realizzate.

Spesso le agenzie web e di marketing si scontrano con problemi legati al blocco dei flussi di dati da parte di sistemi anti-bot (come Cloudflare o Akamai) o perdono ore preziose a formattare manualmente report in Excel per i clienti.

Mi chiamo Vasile e sono uno **Specialista in Python Web Scraping e Automatizzazione Dati**. Aiuto le agenzie a sviluppare motori di scraping personalizzati e resilienti che superano i moderni sistemi di sicurezza (utilizzando Playwright stealth, configurazioni proxy e manipolazione delle impronte digitali TLS) e che generano report pronti per i clienti in formati premium.

Potete visionare alcuni dei miei approfondimenti tecnici direttamente sul mio portfolio: **https://vasiledev.com** o sul mio profilo LinkedIn.

Avete attualmente progetti attivi o clienti che necessitano di estrazione dati su misura o di dashboard automatizzate? Sarei felice di collaborare con voi.

Un cordiale saluto,

Vasile Bratu
Sviluppatore Python & Data Automation
amendamax@vasiledev.com
"""
    elif language == "romanian":
        salute = f"Salut {salutation}" if salutation else "Bună ziua, echipa"
        subject = f"Colaborare Web Scraping și Automatizare Date pentru {company_name}"
        body = f"""{salute} {company_name},

Vă contactez pentru că apreciez portofoliul de proiecte digitale pe care le realizați.

Multe agenții web și de marketing se confruntă des cu probleme de blocare a scripturilor de colectare a datelor de către sisteme anti-bot (precum Cloudflare sau Akamai) sau pierd timp prețios formatând manual rapoarte în Excel pentru clienți.

Numele meu este Vasile și sunt **Specialist în Python Web Scraping și Automatizare**. Ajut agențiile să construiască motoare de scraping personalizate și rezistente, care ocolesc protecțiile moderne (folosind Playwright stealth, proxy-uri rezidențiale și TLS fingerprinting) și livrează datele direct în rapoarte premium, complet formatate.

Puteți vedea câteva dintre analizele mele tehnice direct pe site-ul meu: **https://vasiledev.com** sau pe profilul meu de LinkedIn.

Aveți în acest moment proiecte active sau clienți care au nevoie de extragere de date personalizată sau de rapoarte automatizate? Mi-ar plăcea să colaborăm.

O zi excelentă,

Vasile Bratu
Python & Data Automation Engineer
amendamax@vasiledev.com
"""
    else: # English default
        salute = f"Hi {salutation}" if salutation else "Hi Team"
        subject = f"Custom Python Scrapers & Anti-Bot Bypass for {company_name}"
        body = f"""{salute},

I came across {company_name} and love the digital work you do.

Many web and marketing agencies often face issues with data collection pipelines getting blocked by Cloudflare/Akamai, or they waste hours manually formatting Excel dashboards for their clients.

I’m Vasile, a **Python Web Scraping & Data Automation Specialist**. I help agencies build resilient, custom scraping engines that bypass modern anti-bots (using Playwright, stealth setups, and custom TLS fingerprinting) and output clean, executive-ready reports.

You can check out some of my technical breakdowns on my website: **https://vasiledev.com** or on my LinkedIn profile.

Do you have any active projects or clients requiring custom data extraction or automated reports right now? I would love to help.

Best regards,

Vasile Bratu
Python & Data Automation Specialist
amendamax@vasiledev.com
"""
    return subject, body

def main():
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    excel_path = os.path.join(desktop_dir, "PROSPECTE_AGENTII_OUTREACH.xlsx")
    
    if not os.path.exists(excel_path):
        print(f"[!] Fișierul Excel nu a fost găsit pe Desktop: {excel_path}")
        return
        
    drafts_dir = os.path.join(desktop_dir, "Drafturi_Outreach")
    os.makedirs(drafts_dir, exist_ok=True)
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    email_count = 0
    form_count = 0
    
    form_leads_text = "=== ACCESARE FORMULARE CONTACT (PENTRU CELE FĂRĂ EMAIL DIRECT) ===\n\n"
    
    # Iterate through Excel rows (skip header row 1)
    for row in range(2, ws.max_row + 1):
        company = ws.cell(row=row, column=1).value
        location = ws.cell(row=row, column=2).value
        website = ws.cell(row=row, column=3).value
        email_addr = ws.cell(row=row, column=4).value
        decident = ws.cell(row=row, column=6).value
        source_url = ws.cell(row=row, column=8).value
        
        if not company:
            continue
            
        # Determine language based on location
        location_low = str(location).lower()
        if "italy" in location_low or "milano" in location_low:
            lang = "italian"
        elif "romania" in location_low or "bucurești" in location_low:
            lang = "romanian"
        else:
            lang = "english"
            
        subject, body = get_email_templates(company, decident, lang)
        
        # Scenario A: Email is available -> Generate .eml draft
        if email_addr and email_addr != "N/A" and "exemplu.ro" not in email_addr:
            try:
                msg = EmailMessage()
                msg['Subject'] = subject
                msg['From'] = 'amendamax@vasiledev.com'
                msg['To'] = email_addr
                msg.set_content(body)
                
                safe_company = clean_filename(company)
                eml_path = os.path.join(drafts_dir, f"{safe_company}_Outreach.eml")
                
                with open(eml_path, 'wb') as f:
                    f.write(msg.as_bytes(policy=SMTP))
                email_count += 1
            except Exception as e:
                print(f"[!] Eroare la generarea draftului pentru {company}: {e}")
                
        # Scenario B: No email -> Append to Contact Form helper list
        else:
            form_count += 1
            form_leads_text += f"[{form_count}] {company}\n"
            form_leads_text += f"   Site Web: {website}\n"
            form_leads_text += f"   Pagina Contact/Sursă: {source_url}\n"
            form_leads_text += f"   Subiect propus: {subject}\n"
            form_leads_text += f"   Mesaj pregătit pentru Copy-Paste:\n"
            form_leads_text += f"--------------------------------------------------\n"
            form_leads_text += f"{body}\n"
            form_leads_text += f"==================================================\n\n"
            
    # Save the contact forms helper file
    helper_file_path = os.path.join(drafts_dir, "CONTACT_FORMS_OUTREACH.txt")
    with open(helper_file_path, 'w', encoding='utf-8') as f:
        f.write(form_leads_text)
        
    print(f"\n=======================================================")
    print(f"[+] DRAFTURI GENERATE CU SUCCES PE DESKTOP!")
    print(f"   Folder: Desktop/Drafturi_Outreach")
    print(f"   - E-mailuri (.eml) gata de deschis: {email_count}")
    print(f"   - Firme pentru formular contact (.txt): {form_count}")
    print(f"=======================================================\n")

if __name__ == "__main__":
    main()
