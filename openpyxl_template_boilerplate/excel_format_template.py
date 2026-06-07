from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_styled_excel():
    print("[*] Starting Premium openpyxl Excel Formatting Demo...")
    
    # 1. Create a new workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "B2B Leads Pipeline"
    
    # Ensure grid lines are visible in the spreadsheet
    ws.views.sheetView[0].showGridLines = True
    
    # 2. Define Custom Color Scheme (HSL-inspired "Forest Emerald Green" Theme)
    HEADER_BG = "1B4D3E"       # Deep Emerald Green for headers
    HEADER_FG = "FFFFFF"       # White text for headers
    ZEBRA_BG = "F2F7F5"        # Light mint green zebra fill for readability
    BORDER_COLOR = "E2E8F0"    # Clean, subtle slate gray border
    LINK_COLOR = "1B4D3E"      # Deep emerald green for clickable hyperlinks
    
    # Typography: Segoe UI (Standard across all cells)
    font_header = Font(name="Segoe UI", size=11, bold=True, color=HEADER_FG)
    font_data = Font(name="Segoe UI", size=10, color="000000")
    font_link = Font(name="Segoe UI", size=10, underline="single", color=LINK_COLOR)
    
    # Pattern fills for rows
    fill_header = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_BG, end_color=ZEBRA_BG, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Borders (Thin grid outline)
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    # 3. Create Sample B2B Leads Dataset
    headers = ["Company Name", "Industry", "Location", "Phone Number", "Email Address", "Website Portfolio"]
    leads_data = [
        ["Alpha Tech Solutions", "Software Development", "Milano, Italy", "+39 02 123456", "contact@alphatech.it", "https://alphatech.it"],
        ["Vesta Immobiliare", "Real Estate Agency", "Roma, Italy", "+39 06 987654", "info@vestaimmobiliare.it", "https://vestaimmobiliare.it"],
        ["Apex Manufacturing", "Industrial Supplies", "Torino, Italy", "+39 011 555666", "sales@apexindustrial.it", "https://apexindustrial.it"],
        ["Sleek Web Design", "Marketing Agency", "Milano, Italy", "+39 02 888999", "hello@sleekweb.it", "https://sleekweb.it"],
    ]
    
    # 4. Append Headers and apply styling
    ws.append(headers)
    ws.row_dimensions[1].height = 35  # Vasile's rule: Header row height = 35pt
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    # 5. Append Data Rows and apply layouts
    for row_idx, lead in enumerate(leads_data, start=2):
        ws.row_dimensions[row_idx].height = 32  # Vasile's rule: Data row height = 32pt
        
        company, industry, location, phone, email, website = lead
        
        # Write raw cell values
        ws.cell(row=row_idx, column=1, value=company)
        ws.cell(row=row_idx, column=2, value=industry)
        ws.cell(row=row_idx, column=3, value=location)
        ws.cell(row=row_idx, column=4, value=phone)
        ws.cell(row=row_idx, column=5, value=email)
        
        # Write website column using an interactive HYPERLINK formula to keep sheet compact
        link_formula = f'=HYPERLINK("{website}", "Visit Site ↗")'
        ws.cell(row=row_idx, column=6, value=link_formula)
        
        # Apply formatting cell-by-cell
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            
            # Alignments (Center phone numbers & website links)
            if col_idx in [4, 6]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Fonts (Underline links, standard font for data)
            if col_idx == 6:
                cell.font = font_link
            else:
                cell.font = font_data
                
            # Zebra Striping (Alternate background colors on odd rows)
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
                
    # 6. Dynamic Column Auto-fitting with safety padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            val = str(cell.value or '')
            # For HYPERLINK formulas, count the length of the friendly display name rather than the formula string
            if val.startswith("=HYPERLINK"):
                try:
                    friendly_text = val.split(',')[1].replace('"', '').replace(')', '').strip()
                    max_len = max(max_len, len(friendly_text))
                except IndexError:
                    max_len = max(max_len, len(val))
            else:
                max_len = max(max_len, len(val))
                
        # Set column width with 5 units of extra padding for breathing space
        ws.column_dimensions[col_letter].width = max(max_len + 5, 12)
        
    # 7. Save workbook
    output_filename = "styled_output_demo.xlsx"
    wb.save(output_filename)
    print(f"[+] Styled spreadsheet successfully created at: {output_filename}")

if __name__ == "__main__":
    create_styled_excel()
