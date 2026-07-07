import os
import openpyxl
import glob

search_dirs = [
    r"C:\Users\bratu\Documents\antigravity\amazing-borg",
    r"C:\Users\bratu\Desktop"
]

print("Searching all Excel files for 'viktorija' or 'viktoria'...")
for sdir in search_dirs:
    if not os.path.exists(sdir):
        continue
    xlsx_files = glob.glob(os.path.join(sdir, "*.xlsx")) + glob.glob(os.path.join(sdir, "**", "*.xlsx"), recursive=True)
    for file_path in xlsx_files:
        if "~$" in os.path.basename(file_path):
            continue
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    for c_idx, val in enumerate(row, 1):
                        if val and any(term in str(val).lower() for term in ["viktorija", "viktoria"]):
                            print(f"FOUND in Excel: {file_path} | Sheet: {sheet} | Row: {r_idx}, Col: {c_idx} | Value: {val}")
        except Exception as e:
            # print(f"Error reading {file_path}: {e}")
            pass
print("Finished Excel search.")
