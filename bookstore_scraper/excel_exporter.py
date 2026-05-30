import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
import os
import re

class ExcelReportExporter:
    """
    Exports scraped e-commerce data into a premium, corporate-styled Excel workbook.
    Implements advanced formatting, auto-adjusting column widths, dynamic formulas,
    conditional formatting, and visual charts.
    """

    def __init__(self):
        pass

    def export_data(self, data, output_path):
        """
        Transforms a list of dictionaries into a beautifully styled Excel workbook.
        """
        # 1. Initialize Workbook and Sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scraped Products Report"
        
        # Ensure gridlines are visible (important for professional look)
        ws.views.sheetView[0].showGridLines = True
        
        # Define Columns (using Link representations)
        headers = ["Title", "Price (£)", "Rating (1-5)", "In Stock", "Product Link", "Image Link"]
        
        # 2. Write Header Row
        ws.append(headers)
        
        # 3. Write Data Rows
        for item in data:
            product_url = item.get("Product URL", "")
            image_url = item.get("Image URL", "")
            
            row_data = [
                item.get("Title", ""),
                item.get("Price", 0.0),
                item.get("Rating", 0),
                item.get("In Stock", "No"),
                "View Book ↗" if product_url else "N/A",
                "View Image 📷" if image_url else "N/A"
            ]
            ws.append(row_data)
            
            # Set native Excel hyperlink properties
            current_row = ws.max_row
            if product_url:
                ws.cell(row=current_row, column=5).hyperlink = product_url
            if image_url:
                ws.cell(row=current_row, column=6).hyperlink = image_url

        # 4. Styling Definitions (Steel Blue Palette)
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Dark Steel Blue
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        zebra_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid") # Light grey-blue
        normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border_side = Side(border_style="thin", color="CBD5E0")
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        num_rows = len(data)
        data_start_row = 2
        data_end_row = num_rows + 1

        # 5. Apply Basic Row and Cell Styling
        # Header Row Styling (Height 28)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center if col_idx != 1 else align_left
            cell.border = cell_border

        # Data Rows Styling (Height 32 for spacious airy feel)
        for r_idx in range(data_start_row, data_end_row + 1):
            ws.row_dimensions[r_idx].height = 32
            # Zebra striping
            row_fill = zebra_fill if r_idx % 2 == 0 else normal_fill
            
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.fill = row_fill
                cell.border = cell_border
                
                # Column Specific Formats & Alignments
                if c_idx == 1: # Title
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = align_left
                elif c_idx == 2: # Price
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = align_right
                    cell.number_format = '£#,##0.00'
                elif c_idx == 3: # Rating
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = align_center
                    cell.number_format = '0'
                elif c_idx == 4: # In Stock
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = align_center
                else: # Links (Columns 5 and 6)
                    cell.alignment = align_center
                    if cell.value == "N/A":
                        cell.font = Font(name="Calibri", size=10, italic=True, color="A0AEC0")
                    else:
                        cell.font = Font(name="Calibri", size=10, underline="single", color="1A365D")

        # 6. Add Summary Formulas Row
        summary_row = data_end_row + 1
        ws.row_dimensions[summary_row].height = 24
        
        summary_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Grey highlight
        summary_font = Font(name="Calibri", size=10, bold=True, color="1A365D")
        double_bottom_border = Border(
            top=Side(border_style="thin", color="718096"),
            bottom=Side(border_style="double", color="1A365D"), # Executive double-line accounting border
            left=thin_border_side,
            right=thin_border_side
        )

        # Labels & Formulas
        ws.cell(row=summary_row, column=1, value="Average / Totals").alignment = align_left
        ws.cell(row=summary_row, column=2, value=f"=AVERAGE(B{data_start_row}:B{data_end_row})").number_format = '£#,##0.00'
        ws.cell(row=summary_row, column=3, value=f"=AVERAGE(C{data_start_row}:C{data_end_row})").number_format = '0.0'
        ws.cell(row=summary_row, column=4, value=f'=COUNTIF(D{data_start_row}:D{data_end_row}, "Yes")').number_format = '0'
        ws.cell(row=summary_row, column=5, value="Total Scraped").alignment = align_right
        ws.cell(row=summary_row, column=6, value=f"=COUNTA(A{data_start_row}:A{data_end_row})").number_format = '0'

        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=summary_row, column=c_idx)
            cell.fill = summary_fill
            cell.font = summary_font
            cell.border = double_bottom_border
            if c_idx in [2, 3, 4, 6]:
                cell.alignment = align_center if c_idx in [3, 4] else align_right

        # 7. Apply Conditional Formatting (e.g. green for 5-star ratings, red for 1-star or out of stock)
        green_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
        green_font = Font(color="22543D", bold=True)
        red_fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
        red_font = Font(color="742A2A", bold=True)

        # Rule for 5-star books (Rating == 5)
        ws.conditional_formatting.add(
            f"C{data_start_row}:C{data_end_row}",
            CellIsRule(operator="equal", formula=["5"], fill=green_fill, font=green_font)
        )
        
        # Rule for 1 or 2 star books (low ratings)
        ws.conditional_formatting.add(
            f"C{data_start_row}:C{data_end_row}",
            CellIsRule(operator="lessThanOrEqual", formula=["2"], fill=red_fill, font=red_font)
        )

        # Rule for Out of Stock (In Stock == "No")
        ws.conditional_formatting.add(
            f"D{data_start_row}:D{data_end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=red_fill, font=red_font)
        )

        # 8. Add a Beautiful Price Comparison Chart (first 10 items)
        chart_limit = min(num_rows, 10)
        if chart_limit > 1:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Price Comparison of Scraped Books"
            chart.y_axis.title = "Price (£)"
            chart.x_axis.title = "Book Title"
            
            # Series are columns, Categories are rows (titles)
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=chart_limit + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=chart_limit + 1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.legend = None # No legend needed as there's only 1 series
            
            # Size the chart and position it
            chart.width = 16
            chart.height = 10
            ws.add_chart(chart, "H2") # Position chart at H2

        # 9. Auto-Fit Column Widths (Calculated based on actual display text)
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith('='):
                    val_str = "Average / Totals"
                max_len = max(max_len, len(val_str))
                
            # Restrict title column to maximum 40 width
            if col[0].column == 1:
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 18), 40)
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)



        # Save Workbook
        wb.save(output_path)
        print(f"Excel report saved successfully to: {output_path}")

# Self-test block
if __name__ == "__main__":
    print("Testing Excel Report Exporter...")
    exporter = ExcelReportExporter()
    test_data = [
        {"Title": "A Light in the Attic", "Price": 51.77, "Rating": 3, "In Stock": "Yes", "Product URL": "http://example.com/1", "Image URL": "http://example.com/img1"},
        {"Title": "Tipping the Velvet", "Price": 53.74, "Rating": 1, "In Stock": "Yes", "Product URL": "http://example.com/2", "Image URL": "http://example.com/img2"},
        {"Title": "Soumission", "Price": 50.10, "Rating": 5, "In Stock": "No", "Product URL": "http://example.com/3", "Image URL": "http://example.com/img3"},
        {"Title": "Sharp Objects", "Price": 47.82, "Rating": 4, "In Stock": "Yes", "Product URL": "http://example.com/4", "Image URL": "http://example.com/img4"},
        {"Title": "Sapiens: A Brief History of Humankind", "Price": 54.23, "Rating": 5, "In Stock": "Yes", "Product URL": "http://example.com/5", "Image URL": "http://example.com/img5"},
    ]
    test_file = "test_output.xlsx"
    exporter.export_data(test_data, test_file)
    if os.path.exists(test_file):
        print(f"Self-test success! {test_file} exists.")
        try:
            os.remove(test_file)
            print("Cleanup completed.")
        except Exception:
            pass
