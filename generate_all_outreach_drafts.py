import os
import sys
import glob
import re
import urllib.parse
from email.message import EmailMessage
from email.policy import SMTP
import openpyxl

# Force UTF-8 encoding on Windows console to avoid encoding crashes
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except AttributeError:
    pass


def clean_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "", name).replace(" ", "_")

def get_templates():
    # Real estate Outreach
    re_it_subject = "Automazione dati e reportistica Excel premium per il vostro business"
    re_it_body = """Gentile {owner},

Spero che questa email vi trovi bene.

Sono Vasile Bratu, uno sviluppatore Python senior specializzato in automazione dati e web scraping, residente a Garessio (Cuneo).

Ho analizzato attentamente il settore immobiliare nella zona di {city} e ho notato come la raccolta manuale dei dati di mercato, il monitoraggio degli annunci pubblicati direttamente dai proprietari ("privati") o il caricamento delle schede immobiliari richieda spesso molte ore preziose ogni settimana per il team di {company_name}.

Aiuto le agenzie immobiliari a risparmiare tempo e a battere la concorrenza sul tempo automatizzando questi processi. Nello specifico, posso creare per voi:
1. Estrattori automatici in tempo reale: Per essere sempre i primi a sapere quando un proprietario pubblica un nuovo annuncio sui portali.
2. Dashboard Excel di livello Executive: Report ordinati e spaziosi con formule di hyperlink cliccabili per accedere direttamente agli annunci e alle foto con un solo clic.
3. Sincronizzazione Cloud automatica: Integrazione diretta e sicura con il vostro CRM aziendale o Google Sheets.

Inoltre, ho recentemente pubblicato una guida tecnica approfondita su Medium riguardante la conformità GDPR e lo scraping etico dei dati, che garantisce che le nostre operazioni siano sicure al 100% dal punto di vista legale e tecnico:
👉 https://medium.com/@amendamax/web-scraping-etico-e-gdpr-come-le-aziende-possono-raccogliere-dati-pubblici-online-in-totale-24715b5c76e0

Potete visionare una demo dei miei report premium e i miei progetti open-source sul mio portfolio GitHub qui:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

La mia proposta per voi:
Sarei felice di creare per voi o per il vostro team una demo gratuita con 5 annunci reali della zona di {city}, formattati nel mio database premium, così potrete valutare voi stessi l'utilità del sistema senza alcun impegno.

Fatemi sapere se può interessarvi e quale zona preferite per la demo!

Cordiali saluti,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    re_ro_subject = "Automatizare date si rapoarte Excel inteligente pentru afacerea dumneavoastra"
    re_ro_body = """Buna ziua {owner},

Numele meu este Vasile Bratu si sunt inginer software senior specializat in automatizari de date, web scraping si raportare de business.

Daca echipa dumneavoastra de la {company_name} pierde timp pretios in fiecare zi cu monitorizarea manuala a pietei imobiliare, copierea anunturilor noi postate de proprietari pe diverse portaluri sau curatarea listelor de prospecti din {city}, va pot ajuta sa eliminati complet aceasta munca manuala prin construirea unui pipeline de date automat.

Iata ce pot implementa pentru afacerea dumneavoastra:
1. Scrapere de date in timp real: Extragere automata a anunturilor noi din portalurile imobiliare relevante, direct in secunda in care apar.
2. Rapoarte Excel premium (Executive Dashboards): Livrarea datelor in tabele extrem de aerisite si organizate, cu formule di hyperlink active pentru o navigare extrem de rapida.
3. Sincronizare Cloud: Integrare automata directa in CRM-ul agentiei dumneavoastra sau in Google Sheets.

De asemenea, am publicat recent o analiză pe Medium despre impactul automatizării datelor imobiliare (PropTech în 2026) și cum aceasta elimină munca manuală, crescând eficiența:
👉 https://medium.com/@amendamax/proptech-%C3%AEn-2026-cum-automatizarea-datelor-imobiliare-elimin%C4%83-munca-manual%C4%83-a-agen%C8%9Bilor-%C8%99i-le-e897e6c7b233

Puteti analiza o mostra a calitatii muncii mele si a codului meu pe portofoliul meu GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

Propunerea mea gratuita:
Pentru a va convinge de utilitatea sistemului, sunt bucuros sa realizez un test gratuit cu 5 anunturi reale/date extrase din {city}, formatate in raportul meu premium, fara nicio obligatie din partea dumneavoastra.

Daca vi se pare interesant, va rog sa imi spuneti ce oras/zona va intereseaza pentru testul gratuit!

Cu stima,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    # E-commerce Outreach
    eco_it_subject = "Monitoraggio automatico dei prezzi concorrenti per il vostro e-commerce"
    eco_it_body = """Gentile {owner},

Spero che questa email vi trovi bene.

Sono Vasile Bratu, uno sviluppatore Python senior specializzato in automazione dati e price scraping per e-commerce, residente a Garessio (Cuneo).

Ho analizzato il vostro negozio online {company_name} e ho notato quanto sia cruciale oggi monitorare in tempo reale i prezzi dei concorrenti (su Amazon, eBay o siti web rivali) per rimanere competitivi ed evitare perdite di margine.

Aiuto gli e-commerce di medie dimensioni ad automatizzare il monitoraggio dei prezzi e l'analisi dei cataloghi. Nello specifico, posso creare per voi:
1. Scraping dei prezzi della concorrenza: Monitoraggio automatico e giornaliero dei listini dei vostri concorrenti.
2. Executive Price Dashboard: Report Excel ordinati ed eleganti (in formato "Midnight Gold") con variazioni percentuali, grafici e link rapidi ai prodotti.
3. Riconciliazione automatica nel vostro store: Sincronizzazione dei dati direttamente su Shopify, WooCommerce o Google Sheets.

Inoltre, ho pubblicato una guida tecnica su Medium sulla legalità dello scraping e la conformità al GDPR nell'estrazione di dati pubblici nell'UE:
👉 https://medium.com/@amendamax/web-scraping-etico-e-gdpr-come-le-aziende-possono-raccogliere-dati-pubblici-online-in-totale-24715b5c76e0

Potete visionare una demo dei miei report premium e i miei progetti open-source sul mio portfolio GitHub qui:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

La mia proposta:
Sarei felice di creare una demo gratuita con il monitoraggio in tempo reale di 5 prodotti della concorrenza a vostra scelta, senza alcun impegno.

Fatemi sapere se può interessarvi!

Cordiali saluti,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    eco_ro_subject = "Monitorizare automata preturi concurenta pentru e-commerce"
    eco_ro_body = """Buna ziua {owner},

Numele meu este Vasile Bratu si sunt inginer software senior specializat in automatizari de date si price scraping pentru magazine online.

Daca echipa dumneavoastra de la {company_name} pierde timp pretios in fiecare zi cu monitorizarea manuala a preturilor concurentilor pe diverse site-uri sau pe platforme de tip marketplace, va pot ajuta sa eliminati complet aceasta munca manuala.

Iata ce pot implementa pentru magazinul dumneavoastra online:
1. Scrapere de preturi in timp real: Urmarirea automata a schimbarilor de pret de pe site-urile concurente sau marketplace-uri.
2. Rapoarte Excel premium (Midnight Gold Dashboards): Tabele aerisite si organizate cu semnalari vizuale pentru oportunitatile de crestere a pretului sau reduceri de pret necesare.
3. Sincronizare automata: Integrare directa in platforma dumneavoastra e-commerce (Shopify/WooCommerce) sau in Google Sheets.

De asemenea, am publicat recent un articol detaliat pe Medium despre inteligența prețurilor în e-commerce și modul în care urmărirea automată a concurenței vă protejează marja de profit:
👉 https://medium.com/@amendamax/inteligen%C8%9Ba-pre%C8%9Burilor-%C3%AEn-e-commerce-cum-automatizarea-monitoriz%C4%83rii-competitorilor-protejeaz%C4%83-7abde4e1e508

Puteti analiza o mostra a calitatii muncii mele si a codului meu pe portofoliul meu GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

Propunerea mea gratuita:
Sunt bucuros sa realizez un test gratuit cu monitorizarea automata a 5 produse din magazinul dumneavoastra in raport cu concurenta, fara nicio obligatie.

Daca vi se pare interesant, va rog sa imi spuneti ce produse doriti sa monitorizam!

Cu stima,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    # English Real Estate template
    re_en_subject = "Automating your property data pipeline & custom Excel reporting"
    re_en_body = """Hi {owner},

I hope this email finds you well.

I'm Vasile Bratu, a Senior Python & Data Automation Engineer specializing in web scraping, API integration, and executive-ready reporting dashboards.

I analyzed the real estate market in your area and noticed how much manual time teams at {company_name} spend copy-pasting property listings, monitoring private seller ads, or updating internal listings in {city}.

I help real estate agencies save time and beat competitors by automating these workflows. Specifically, I can build for you:
1. Real-time property scrapers: To be the first to know when a private owner lists a new property.
2. Executive Excel Dashboards: Spacious, beautiful reports (Midnight Gold & Forest Green) with active, clickable hyperlink formulas for one-click access.
3. CRM Sync: Real-time synchronization directly into Google Sheets, HubSpot, or Salesforce APIs.

Additionally, I recently published a technical breakdown on Dev.to about FinTech, API integration, and VPS automated risk control, which demonstrates my approach to high-performance and resilient systems engineering:
👉 https://dev.to/amendamax2025/fintech-algorithmic-risk-control-how-vps-automation-and-api-integration-protect-capital-and-25c8

You can review my open-source code and portfolio repositories on GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

My risk-free offer:
I am ready to perform a 5-lead free trial targeting your preferred area formatted in my premium report so you can evaluate the speed and quality firsthand.

Let me know if this sounds interesting!

Best regards,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    # English E-commerce template
    eco_en_subject = "Automating your competitor price monitoring & e-commerce reporting"
    eco_en_body = """Hi {owner},

I hope this email finds you well.

I'm Vasile Bratu, a Senior Python & Data Automation Engineer specializing in web scraping, API integration, and competitor price tracking for e-commerce stores.

I analyzed your online store {company_name} and noticed how crucial real-time competitor price monitoring (on Amazon, eBay, or direct competitor sites) is today to stay competitive and protect your margins.

I help e-commerce brands automate price tracking and catalog analysis. Specifically, I can build for you:
1. High-Performance Competitor Price Scraping: Automated daily tracking of your competitors' prices.
2. Executive Price Dashboards: Breathtaking reports (Midnight Gold & Forest Green) with percentage variations, price history charts, and direct product links.
3. CRM/Store Sync: Real-time synchronization directly into Shopify, WooCommerce, or Google Sheets.

Additionally, I recently published a detailed technical guide on Dev.to regarding ethical web scraping and GDPR compliance, which ensures your data operations are 100% legally and technically secure:
👉 https://dev.to/amendamax2025/ethical-web-scraping-gdpr-how-enterprises-extract-public-web-data-with-absolute-legal--1fb9

You can review my open-source code and portfolio repositories on GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

My risk-free offer:
I am ready to perform a 5-product free competitor price tracking trial for you so you can see the results firsthand.

Let me know if this sounds interesting!

Best regards,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    # Web Agency
    wa_it_subject = "Supporto tecnico Python / Web Scraping per i progetti di {company_name}"
    wa_it_body = """Gentile {owner},

Spero che questa email vi trovi bene.

Sono Vasile Bratu, uno sviluppatore Python senior specializzato in Web Scraping, automazione dati ed estrazioni complesse, residente in Italia (Cuneo).

Nel corso dei miei progetti, collaboro spesso con agenzie web e di digital marketing come {company_name} per supportare lo sviluppo tecnico di soluzioni che richiedono la raccolta automatizzata di dati, l'integrazione di API personalizzate o il superamento di barriere anti-bot complesse (Cloudflare, Akamai, Datadome).

Se il vostro team si trova a dover gestire richieste di scraping di dati, migrazioni complesse di database e-commerce, monitoraggio costante dei prezzi per conto dei vostri clienti o reporting automatizzato in Excel/Google Sheets, posso offrirvi un supporto esterno rapido e professionale per:
1. Sviluppo di scraper custom robusti (Scrapy, Playwright, Selenium).
2. Bypass di protezioni anti-bot avanzate tramite tecniche di impersonazione e proxy rotation.
3. Generazione di report Excel di livello executive direttamente via script (con libreria openpyxl).

Ho recentemente pubblicato due guide tecniche di rilievo per sviluppatori ed agenzie su Medium:
- Come bypassare Cloudflare & WAF in modo etico e sicuro:
  👉 https://medium.com/@amendamax/web-scraping-etico-e-gdpr-come-le-aziende-possono-raccogliere-dati-pubblici-online-in-totale-24715b5c76e0
- Generazione di report Excel professionali con Python:
  👉 https://medium.com/@amendamax/generazione-di-report-excel-premium-in-python-con-openpyxl-una-guida-per-sviluppatori-c44bbff633ee

Potete visionare i miei scraper open-source ed esempi di report sul mio profilo GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

La mia proposta per voi:
Offro la mia disponibilità per realizzare un piccolo test/demo gratuito di scraping su un sito target a vostra scelta (ad esempio, estrarre i primi dati da un portale che i vostri clienti monitorano), per dimostrarvi la qualità del codice e dei dati estratti senza alcun impegno da parte vostra.

Sarei felice di fare una breve chiamata conoscitiva se ritenete che possa nascere una collaborazione.

Un cordiale saluto,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub: https://github.com/amendamax"""

    wa_ro_subject = "Colaborare tehnica Python / Web Scraping pentru proiectele {company_name}"
    wa_ro_body = """Buna ziua {owner},

Numele meu este Vasile Bratu si sunt inginer software senior specializat in Web Scraping, automatizari de date si integrari de sisteme in Python.

Colaborez frecvent cu agentii web si agentii de digital marketing pentru a le ajuta sa externalizeze sarcini tehnice complexe legate de extragerea automatizata a datelor, migrarea bazelor de date, monitorizarea concurentei pentru clientii lor sau bypass-ul protectiilor anti-bot (Cloudflare, Akamai).

Va pot sustine echipa de la {company_name} ca partener tehnic extern pentru:
1. Dezvoltarea de scraper-e custom, rezistente la blocaje (Scrapy, Playwright).
2. Automatizarea rapoartelor de business direct in format Excel premium (folosind openpyxl).
3. Integrarea API-urilor si sincronizarea datelor in Cloud/CRM.

Portofoliul meu open-source si mostre de rapoarte pot fi consultate pe GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

Propunerea mea gratuita:
Sunt bucuros sa realizez un test/demo gratuit (extragerea catorva date dintr-un site target ales de dumneavoastra), pentru a va convinge de calitatea datelor livrate si a codului meu, fara nicio obligatie.

Daca doriti o scurta discutie sau un test gratuit, va rog sa imi lasati un mesaj.

Cu stima,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub: https://github.com/amendamax"""

    wa_en_subject = "Python / Web Scraping technical support for {company_name} projects"
    wa_en_body = """Hi {owner},

Hope you are doing well.

I am Vasile Bratu, a senior Python developer specializing in web scraping, data pipeline automation, and anti-bot bypass (Cloudflare, Akamai).

I frequently collaborate with web development and SEO agencies like {company_name} as an external technical partner, handling complex web scraping, database migrations, competitor price monitoring for their clients, and automated Excel reporting.

I can support your team with:
1. Development of robust, production-grade custom scrapers (Scrapy, Playwright).
2. Advanced anti-bot bypass strategies (impersonation, proxy rotation).
3. Automating high-end Excel executive reports via Python scripting.

I recently published a couple of technical articles on Medium regarding GDPR compliance in scraping and professional Excel formatting:
- Web scraping ethics & GDPR:
  👉 https://medium.com/@amendamax/web-scraping-etico-e-gdpr-come-le-aziende-possono-raccogliere-dati-pubblici-online-in-totale-24715b5c76e0
- Generating premium Excel reports with Openpyxl:
  👉 https://medium.com/@amendamax/generazione-di-report-excel-premium-in-python-con-openpyxl-una-guida-per-sviluppatori-c44bbff633ee

You can review my open-source scrapers and report samples on my GitHub profile:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

My free offer for you:
I'm happy to write a free proof-of-concept (POC) script to scrape a small sample from a target website of your choice, so you can evaluate the quality of the data and delivery with zero commitments.

Let me know if you would be open to a quick introductory call!

Best regards,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub: https://github.com/amendamax"""

    return {
        "Imobiliare": {
            "italy": (re_it_subject, re_it_body),
            "romania": (re_ro_subject, re_ro_body),
            "usa": (re_en_subject, re_en_body),
            "uk": (re_en_subject, re_en_body)
        },
        "E-commerce": {
            "italy": (eco_it_subject, eco_it_body),
            "romania": (eco_ro_subject, eco_ro_body),
            "usa": (eco_en_subject, eco_en_body),
            "uk": (eco_en_subject, eco_en_body)
        },
        "Web Agency": {
            "italy": (wa_it_subject, wa_it_body),
            "romania": (wa_ro_subject, wa_ro_body),
            "usa": (wa_en_subject, wa_en_body),
            "uk": (wa_en_subject, wa_en_body)
        }
    }

def main():
    workspace_dir = r"C:\Users\bratu\Documents\antigravity\amazing-borg"
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    drafts_dir = os.path.join(desktop_dir, "Drafturi_Outreach")
    
    # Clean the drafts directory first
    if os.path.exists(drafts_dir):
        for f_old in glob.glob(os.path.join(drafts_dir, "*")):
            try:
                os.remove(f_old)
            except Exception as e:
                pass
    else:
        os.makedirs(drafts_dir, exist_ok=True)

    templates = get_templates()
    
    excel_files = glob.glob(os.path.join(workspace_dir, "leads_*.xlsx"))
    outreach_file = os.path.join(workspace_dir, "PROSPECTE_AGENTII_OUTREACH.xlsx")
    if os.path.exists(outreach_file):
        excel_files.append(outreach_file)
        
    email_count = 0
    form_count = 0
    form_leads_text = "=== ACCESARE FORMULARE CONTACT (FĂRĂ EMAIL DIRECT) ===\n\n"
    
    print(f"Scanning workspace for leads spreadsheets. Found {len(excel_files)} files.")
    
    for file_path in excel_files:
        filename = os.path.basename(file_path).lower()
        if "contacted_success" in filename:
            continue
            
        # Niche determination
        if "ecommerce" in filename:
            niche = "E-commerce"
        elif "webagency" in filename or "agency" in filename or "prospecte_agentii_outreach" in filename:
            niche = "Web Agency"
        else:
            niche = "Imobiliare"
            
        # City and Country determination (defaults based on filename)
        if "torino" in filename:
            city, country = "Torino", "italy"
        elif "milano" in filename:
            city, country = "Milano", "italy"
        elif "bucuresti" in filename:
            city, country = "București", "romania"
        elif "roma" in filename:
            city, country = "Roma", "italy"
        elif "london" in filename:
            city, country = "London", "uk"
        elif "new york" in filename:
            city, country = "New York", "usa"
        elif "cluj" in filename:
            city, country = "Cluj", "romania"
        elif "miami" in filename:
            city, country = "Miami", "usa"
        elif "iasi" in filename:
            city, country = "Iași", "romania"
        elif "timisoara" in filename:
            city, country = "Timișoara", "romania"
        elif "manchester" in filename:
            city, country = "Manchester", "uk"
        else:
            city, country = "Global", "usa"
            
        print(f"  Processing file: {os.path.basename(file_path)} (Niche: {niche}, City: {city}, Country: {country})")
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                company = ws.cell(row=row, column=1).value
                if not company:
                    continue
                    
                cell_web = ws.cell(row=row, column=4)
                website = cell_web.hyperlink.target if cell_web.hyperlink else cell_web.value
                email_addr = ws.cell(row=row, column=7).value if ws.max_column >= 7 else ws.cell(row=row, column=4).value
                owner = ws.cell(row=row, column=6).value if ws.max_column >= 6 else "N/A"
                
                # Double-check column indices
                if row == 2:
                    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                    email_col_idx = None
                    owner_col_idx = None
                    web_col_idx = None
                    for c_idx, h in enumerate(headers, start=1):
                        h_str = str(h).lower()
                        if "email" in h_str or "e-mail" in h_str:
                            email_col_idx = c_idx
                        elif "decident" in h_str or "proprietar" in h_str or "owner" in h_str:
                            owner_col_idx = c_idx
                        elif "site" in h_str or "website" in h_str:
                            web_col_idx = c_idx
                
                if email_col_idx:
                    email_addr = ws.cell(row=row, column=email_col_idx).value
                if owner_col_idx:
                    owner = ws.cell(row=row, column=owner_col_idx).value
                if web_col_idx:
                    cell_web = ws.cell(row=row, column=web_col_idx)
                    website = cell_web.hyperlink.target if cell_web.hyperlink else cell_web.value

                # Determine city/country dynamically for row if it's the general outreach list
                row_city, row_country = city, country
                if "prospecte_agentii_outreach" in filename:
                    loc_val = str(ws.cell(row=row, column=2).value).lower()
                    if "milano" in loc_val or "italy" in loc_val:
                        row_city, row_country = "Milano", "italy"
                    elif "bucuresti" in loc_val or "romania" in loc_val or "bucurești" in loc_val:
                        row_city, row_country = "București", "romania"
                    elif "london" in loc_val or "uk" in loc_val:
                        row_city, row_country = "London", "uk"
                    elif "new york" in loc_val or "usa" in loc_val:
                        row_city, row_country = "New York", "usa"
                    else:
                        row_city, row_country = "Global", "usa"

                # Filter out placeholder emails
                email_str = str(email_addr).strip()
                ignore_domains = ['example.com', 'domain.com', 'yourcompany.co.uk', 'yourdomain.com', 'yourdomain', 'example']
                is_placeholder = any(ig in email_str.lower() for ig in ignore_domains)
                if is_placeholder:
                    continue

                # Determine salutation
                owner_str = str(owner).strip()
                fake_owners = [
                    "mario rossi", "giuseppe bianchi", "andrei popescu", "mihai ionescu", "john smith", "david davis", 
                    "google", "marketing", "google review", "of mt salons", "to collect all the inform", 
                    "colaborare bun", "to collect all the info"
                ]
                
                is_owner_fake = (
                    owner_str.lower() in fake_owners or 
                    owner_str.lower() == "none" or 
                    owner_str.lower() == "n/a" or 
                    len(owner_str) < 3 or 
                    any(w in owner_str.lower() for w in ["review", "salon", "collect", "review count"])
                )
                
                if not is_owner_fake:
                    # Use first name of owner if a person
                    owner_val = owner_str.split()[0]
                else:
                    # Default placeholders
                    if row_country == "italy":
                        owner_val = "Responsabile Tecnico" if niche == "Web Agency" else "Responsabile"
                    elif row_country == "romania":
                        owner_val = "Responsabil Tehnic" if niche == "Web Agency" else f"Echipa {company}"
                    else:
                        owner_val = "Technical Lead" if niche == "Web Agency" else f"Team at {company}"

                subject, body = templates[niche][row_country]
                formatted_subj = subject
                formatted_body = body.format(
                    owner=owner_val,
                    city=row_city,
                    company_name=company
                )
                
                # Check email validity
                is_email_valid = email_addr and email_str != "N/A" and email_str != "None" and "@" in email_str
                
                if is_email_valid:
                    try:
                        msg = EmailMessage()
                        msg['Subject'] = formatted_subj
                        msg['From'] = 'amendamax@vasiledev.com'
                        msg['To'] = email_str
                        msg.set_content(formatted_body)
                        
                        safe_niche = clean_filename(niche)
                        safe_city = clean_filename(row_city)
                        safe_company = clean_filename(company)
                        eml_filename = f"{safe_niche}_{safe_city}_{safe_company}_Outreach.eml"
                        eml_path = os.path.join(drafts_dir, eml_filename)
                        
                        with open(eml_path, 'wb') as f:
                            f.write(msg.as_bytes(policy=SMTP))
                        email_count += 1
                    except Exception as e:
                        print(f"    [!] Error writing EML for {company}: {e}")
                else:
                    form_count += 1
                    form_leads_text += f"[{form_count}] {company} ({niche} - {row_city}, {row_country.upper()})\n"
                    form_leads_text += f"   Website: {website}\n"
                    form_leads_text += f"   Suggested Subject: {formatted_subj}\n"
                    form_leads_text += f"   Outreach Message:\n"
                    form_leads_text += f"--------------------------------------------------\n"
                    form_leads_text += f"{formatted_body}\n"
                    form_leads_text += f"==================================================\n\n"
                    
        except Exception as e:
            print(f"  [!] Error processing file {file_path}: {e}")

    # Write the forms contact file
    helper_path = os.path.join(drafts_dir, "CONTACT_FORMS_OUTREACH.txt")
    with open(helper_path, "w", encoding="utf-8") as f:
        f.write(form_leads_text)
        
    print(f"\n=======================================================")
    print(f"[+] DRAFTING COMPLETED SUCCESSFULLY!")
    print(f"    Drafts saved to Desktop/Drafturi_Outreach/")
    print(f"    - Email drafts (.eml) generated: {email_count}")
    print(f"    - Contact form helpers generated: {form_count}")
    print(f"=======================================================\n")
if __name__ == "__main__":
    main()
