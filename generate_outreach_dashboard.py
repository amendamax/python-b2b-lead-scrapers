import os
import urllib.parse
import openpyxl

def get_leads_from_excel(file_path, niche):
    if not os.path.exists(file_path):
        return []
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    leads = []
    
    file_name_lower = os.path.basename(file_path).lower()
    is_outreach_file = "prospecte_agentii_outreach" in file_name_lower
    
    # Read rows starting from row 2 (skipping header)
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        # If row is empty, skip
        if not name:
            continue
            
        if is_outreach_file:
            # Col 1: Name, Col 2: Location/City/Country, Col 3: Website, Col 4: Email, Col 5: Phone, Col 6: Owner
            address = ws.cell(row=row, column=2).value
            cell_web = ws.cell(row=row, column=3)
            website = cell_web.hyperlink.target if cell_web.hyperlink else cell_web.value
            email = ws.cell(row=row, column=4).value
            phone = ws.cell(row=row, column=5).value
            owner = ws.cell(row=row, column=6).value
        else:
            category = ws.cell(row=row, column=2).value
            address = ws.cell(row=row, column=3).value
            # Reconstruct website URL from hyperlink if possible
            cell_web = ws.cell(row=row, column=4)
            website = cell_web.hyperlink.target if cell_web.hyperlink else cell_web.value
            phone = ws.cell(row=row, column=5).value
            owner = ws.cell(row=row, column=6).value
            email = ws.cell(row=row, column=7).value
            
        # Filter out placeholder emails
        email_str = str(email).strip()
        ignore_domains = ['example.com', 'domain.com', 'yourcompany.co.uk', 'yourdomain.com', 'yourdomain', 'example']
        is_placeholder = any(ig in email_str.lower() for ig in ignore_domains)
        if is_placeholder or not email_str or email_str == "N/A" or email_str == "None" or "@" not in email_str:
            continue

        # Clean fake owners
        owner_str = str(owner).strip()
        fake_owners = [
            "mario rossi", "giuseppe bianchi", "andrei popescu", "mihai ionescu", "john smith", "david davis", 
            "google", "marketing", "google review", "of mt salons", "to collect all the inform", 
            "colaborare bun", "to collect all the info"
        ]
        is_owner_fake = (
            owner_str.lower() in fake_owners or 
            owner_str.lower() == "none" or 
            owner_str.lower() == "n/a" or 
            len(owner_str) < 3 or 
            any(w in owner_str.lower() for w in ["review", "salon", "collect", "review count"])
        )
        if is_owner_fake:
            owner = "None"

        # Deduce city and country dynamically
        if is_outreach_file:
            loc_val = str(address).lower()
            if "milano" in loc_val or "italy" in loc_val:
                city = "Milano"
                country = "italy"
            elif "bucuresti" in loc_val or "romania" in loc_val or "bucurești" in loc_val:
                city = "București"
                country = "romania"
            elif "london" in loc_val or "uk" in loc_val:
                city = "London"
                country = "uk"
            elif "new york" in loc_val or "usa" in loc_val:
                city = "New York"
                country = "usa"
            else:
                city = "Global"
                country = "usa"
        else:
            if "torino" in file_name_lower:
                city = "Torino"
                country = "italy"
            elif "milano" in file_name_lower:
                city = "Milano"
                country = "italy"
            elif "bucuresti" in file_name_lower:
                city = "București"
                country = "romania"
            elif "roma" in file_name_lower:
                city = "Roma"
                country = "italy"
            elif "london" in file_name_lower:
                city = "London"
                country = "uk"
            elif "new york" in file_name_lower:
                city = "New York"
                country = "usa"
            elif "cluj" in file_name_lower:
                city = "Cluj"
                country = "romania"
            elif "iasi" in file_name_lower:
                city = "Iași"
                country = "romania"
            elif "timisoara" in file_name_lower:
                city = "Timișoara"
                country = "romania"
            elif "manchester" in file_name_lower:
                city = "Manchester"
                country = "uk"
            elif "miami" in file_name_lower:
                city = "Miami"
                country = "usa"
            else:
                city = "Global"
                country = "usa"
        
        leads.append({
            "company_name": name,
            "niche": niche,
            "city": city,
            "country": country,
            "website": website,
            "phone": phone,
            "owner": owner,
            "email": email
        })
    return leads
def get_templates():
    # Real estate Outreach
    re_it_subject = "Automazione dati e reportistica Excel premium per il vostro business"
    re_it_body = """Gentile {owner},

Spero che questa email vi trovi bene.

Sono Vasile Bratu, uno sviluppatore Python senior specializzato in automazione dati e web scraping, residente a Garessio (Cuneo).

Ho analizzato attentamente il settore immobiliare nella zona di {city} e ho notato come la raccolta manuale dei dati di mercato, il monitoraggio degli annunci pubblicati direttamente dai proprietari ("privati") o il caricamento delle schede immobiliari richieda spesso molte ore preziose ogni settimana per il team di {company_name}.

Aiuto le agenzie immobiliari a risparmiare tempo e a battere la concorrenza sul tempo automatizzando questi processi. Nello specifico, posso creare per voi:
1. Estrattori automatici in tempo reale: Per essere sempre i primi a sapere quando un proprietario pubblica un nuovo annuncio sui portali.
2. Dashboard Excel di livello Executive: Report ordinati e spaziosi con formule di hyperlink cliccabili per accedere direttamente agli annunci e alle foto con un solo clic.
3. Sincronizzazione Cloud automatica: Integrazione diretta e sicura con il vostro CRM aziendale o Google Sheets.

Inoltre, ho recentemente pubblicato una guida tecnica approfondita su Medium riguardante la conformità GDPR e lo scraping etico dei dati, che garantisce che le nostre operazioni siano sicure al 100% dal punto di vista legale e tecnico:
👉 https://medium.com/@amendamax/web-scraping-etico-e-gdpr-come-le-aziende-possono-raccogliere-dati-pubblici-online-in-totale-24715b5c76e0

Potete visionare una demo dei miei report premium e i miei progetti open-source sul mio portfolio GitHub qui:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

La mia proposta per voi:
Sarei felice di creare per voi o per il vostro team una demo gratuita con 5 annunci reali della zona di {city}, formattati nel mio database premium, così potrete valutare voi stessi l'utilità del sistema senza alcun impegno.

Fatemi sapere se può interessarvi e quale zona preferite per la demo!

Cordiali saluti,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    re_ro_subject = "Automatizare date si rapoarte Excel inteligente pentru afacerea dumneavoastra"
    re_ro_body = """Buna ziua {owner},

Numele meu este Vasile Bratu si sunt inginer software senior specializat in automatizari de date, web scraping si raportare de business.

Daca echipa dumneavoastra de la {company_name} pierde timp pretios in fiecare zi cu monitorizarea manuala a pietei imobiliare, copierea anunturilor noi postate de proprietari pe diverse portaluri sau curatarea listelor de prospecti din {city}, va pot ajuta sa eliminati complet aceasta munca manuala prin construirea unui pipeline de date automat.

Iata ce pot implementa pentru afacerea dumneavoastra:
1. Scrapere de date in timp real: Extragere automata a anunturilor noi din portalurile imobiliare relevante, direct in secunda in care apar.
2. Rapoarte Excel premium (Executive Dashboards): Livrarea datelor in tabele extrem de aerisite si organizate, cu formule di hyperlink active pentru o navigare extrem de rapida.
3. Sincronizare Cloud: Integrare automata directa in CRM-ul agentiei dumneavoastra sau in Google Sheets.

De asemenea, am publicat recent o analiză pe Medium despre impactul automatizării datelor imobiliare (PropTech în 2026) și cum aceasta elimină munca manuală, crescând eficiența:
👉 https://medium.com/@amendamax/proptech-%C3%AEn-2026-cum-automatizarea-datelor-imobiliare-elimin%C4%83-munca-manual%C4%83-a-agen%C8%9Bilor-%C8%99i-le-e897e6c7b233

Puteti analiza o mostra a calitatii muncii mele si a codului meu pe portofoliul meu GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

Propunerea mea gratuita:
Pentru a va convinge de utilitatea sistemului, sunt bucuros sa realizez un test gratuit cu 5 anunturi reale/date extrase din {city}, formatate in raportul meu premium, fara nicio obligatie din partea dumneavoastra.

Daca vi se pare interesant, va rog sa imi spuneti ce oras/zona va intereseaza pentru testul gratuit!

Cu stima,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    # E-commerce Outreach
    eco_it_subject = "Monitoraggio automatico dei prezzi concorrenti per il vostro e-commerce"
    eco_it_body = """Gentile {owner},

Spero che questa email vi trovi bene.

Sono Vasile Bratu, uno sviluppatore Python senior specializzato in automazione dati e price scraping per e-commerce, residente a Garessio (Cuneo).

Ho analizzato il vostro negozio online {company_name} e ho notato quanto sia cruciale oggi monitorare in tempo reale i prezzi dei concorrenti (su Amazon, eBay o siti web rivali) per rimanere competitivi ed evitare perdite di margine.

Aiuto gli e-commerce di medie dimensioni ad automatizzare il monitoraggio dei prezzi e l'analisi dei cataloghi. Nello specifico, posso creare per voi:
1. Scraping dei prezzi della concorrenza: Monitoraggio automatico e giornaliero dei listini dei vostri concorrenti.
2. Executive Price Dashboard: Report Excel ordinati ed eleganti (in formato "Midnight Gold") con variazioni percentuali, grafici e link rapidi ai prodotti.
3. Riconciliazione automatica nel vostro store: Sincronizzazione dei dati direttamente su Shopify, WooCommerce o Google Sheets.

Inoltre, ho pubblicato una guida tecnica su Medium sulla legalità dello scraping e la conformità al GDPR nell'estrazione di dati pubblici nell'UE:
👉 https://medium.com/@amendamax/web-scraping-etico-e-gdpr-come-le-aziende-possono-raccogliere-dati-pubblici-online-in-totale-24715b5c76e0

Potete visionare una demo dei miei report premium e i miei progetti open-source sul mio portfolio GitHub qui:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

La mia proposta:
Sarei felice di creare una demo gratuita con il monitoraggio in tempo reale di 5 prodotti della concorrenza a vostra scelta, senza alcun impegno.

Fatemi sapere se può interessarvi!

Cordiali saluti,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    eco_ro_subject = "Monitorizare automata preturi concurenta pentru e-commerce"
    eco_ro_body = """Buna ziua {owner},

Numele meu este Vasile Bratu si sunt inginer software senior specializat in automatizari de date si price scraping pentru magazine online.

Daca echipa dumneavoastra de la {company_name} pierde timp pretios in fiecare zi cu monitorizarea manuala a preturilor concurentilor pe diverse site-uri sau pe platforme de tip marketplace, va pot ajuta sa eliminati complet aceasta munca manuala.

Iata ce pot implementa pentru magazinul dumneavoastra online:
1. Scrapere de preturi in timp real: Urmarirea automata a schimbarilor de pret de pe site-urile concurente sau marketplace-uri.
2. Rapoarte Excel premium (Midnight Gold Dashboards): Tabele aerisite si organizate cu semnalari vizuale pentru oportunitatile de crestere a pretului sau reduceri de pret necesare.
3. Sincronizare automata: Integrare directa in platforma dumneavoastra e-commerce (Shopify/WooCommerce) sau in Google Sheets.

De asemenea, am publicat recent un articol detaliat pe Medium despre inteligența prețurilor în e-commerce și modul în care urmărirea automată a concurenței vă protejează marja de profit:
👉 https://medium.com/@amendamax/inteligen%C8%9Ba-pre%C8%9Burilor-%C3%AEn-e-commerce-cum-automatizarea-monitoriz%C4%83rii-competitorilor-protejeaz%C4%83-7abde4e1e508

Puteti analiza o mostra a calitatii muncii mele si a codului meu pe portofoliul meu GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

Propunerea mea gratuita:
Sunt bucuros sa realizez un test gratuit cu monitorizarea automata a 5 produse din magazinul dumneavoastra in raport cu concurenta, fara nicio obligatie.

Daca vi se pare interesant, va rog sa imi spuneti ce produse doriti sa monitorizam!

Cu stima,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    # English Real Estate template
    re_en_subject = "Automating your property data pipeline & custom Excel reporting"
    re_en_body = """Hi {owner},

I hope this email finds you well.

I'm Vasile Bratu, a Senior Python & Data Automation Engineer specializing in web scraping, API integration, and executive-ready reporting dashboards.

I analyzed the real estate market in your area and noticed how much manual time teams at {company_name} spend copy-pasting property listings, monitoring private seller ads, or updating internal listings in {city}.

I help real estate agencies save time and beat competitors by automating these workflows. Specifically, I can build for you:
1. Real-time property scrapers: To be the first to know when a private owner lists a new property.
2. Executive Excel Dashboards: Spacious, beautiful reports (Midnight Gold & Forest Green) with active, clickable hyperlink formulas for one-click access.
3. CRM Sync: Real-time synchronization directly into Google Sheets, HubSpot, or Salesforce APIs.

Additionally, I recently published a technical breakdown on Dev.to about FinTech, API integration, and VPS automated risk control, which demonstrates my approach to high-performance and resilient systems engineering:
👉 https://dev.to/amendamax2025/fintech-algorithmic-risk-control-how-vps-automation-and-api-integration-protect-capital-and-25c8

You can review my open-source code and portfolio repositories on GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

My risk-free offer:
I am ready to perform a 5-lead free trial targeting your preferred area formatted in my premium report so you can evaluate the speed and quality firsthand.

Let me know if this sounds interesting!

Best regards,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    # English E-commerce template
    eco_en_subject = "Automating your competitor price monitoring & e-commerce reporting"
    eco_en_body = """Hi {owner},

I hope this email finds you well.

I'm Vasile Bratu, a Senior Python & Data Automation Engineer specializing in web scraping, API integration, and competitor price tracking for e-commerce stores.

I analyzed your online store {company_name} and noticed how crucial real-time competitor price monitoring (on Amazon, eBay, or direct competitor sites) is today to stay competitive and protect your margins.

I help e-commerce brands automate price tracking and catalog analysis. Specifically, I can build for you:
1. High-Performance Competitor Price Scraping: Automated daily tracking of your competitors' prices.
2. Executive Price Dashboards: Breathtaking reports (Midnight Gold & Forest Green) with percentage variations, price history charts, and direct product links.
3. CRM/Store Sync: Real-time synchronization directly into Shopify, WooCommerce, or Google Sheets.

Additionally, I recently published a detailed technical guide on Dev.to regarding ethical web scraping and GDPR compliance, which ensures your data operations are 100% legally and technically secure:
👉 https://dev.to/amendamax2025/ethical-web-scraping-gdpr-how-enterprises-extract-public-web-data-with-absolute-legal--1fb9

You can review my open-source code and portfolio repositories on GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

My risk-free offer:
I am ready to perform a 5-product free competitor price tracking trial for you so you can see the results firsthand.

Let me know if this sounds interesting!

Best regards,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub Portfolio: https://github.com/amendamax"""

    # === FOLLOW-UP TEMPLATES ===
    # Real estate Follow-up
    re_it_followup_subj = "Re: Automazione dati e reportistica Excel premium per il vostro business"
    re_it_followup_body = """Gentile {owner},

Le scrivo per un breve follow-up al mio messaggio precedente. Mi scuso per l'invio da un indirizzo diverso; ho recentemente migrato le mie comunicazioni sulla mia email aziendale ufficiale (amendamax@vasiledev.com) e volevo assicurarmi che il mio messaggio precedente fosse arrivato.

Se il team di {company_name} desidera ancora automatizzare il monitoraggio degli annunci o la raccolta dei dati di mercato in {city}, la mia proposta per una demo gratuita con 5 dati reali rimane valida.

Se le interessa vedere come strutturo questi dati, può leggere il mio ultimo articolo su come generare report Excel premium pronti per i clienti direttamente in Python:
👉 https://dev.to/amendamax2025/how-to-build-executive-ready-excel-reports-directly-in-python-using-openpyxl-393c

Mi faccia sapere se desidera che prepari la demo gratuita!

Cordiali saluti,

Vasile Bratu
https://vasiledev.com
amendamax@vasiledev.com"""

    re_ro_followup_subj = "Re: Automatizare date si rapoarte Excel inteligente pentru afacerea dumneavoastra"
    re_ro_followup_body = """Bună ziua {owner},

Revin cu un scurt follow-up la mesajul meu anterior. Scuze pentru trimiterea de pe o adresă diferită; tocmai am migrat comunicarea pe adresa mea profesională oficială (amendamax@vasiledev.com) și am vrut să mă asigur că mesajul meu a ajuns cu bine la dumneavoastră.

Dacă echipa {company_name} dorește în continuare să elimine munca manuală cu monitorizarea proprietăților din {city}, rămâne valabilă oferta mea de a crea un test/demo gratuit cu 5 date reale. 

De asemenea, am publicat recent un ghid tehnic despre cum generez aceste rapoarte Excel premium (design, grafice și structură) pe care îl puteți parcurge aici:
👉 https://dev.to/amendamax2025/how-to-build-executive-ready-excel-reports-directly-in-python-using-openpyxl-393c

Spuneți-mi dacă v-ar interesa acest test gratuit!

Cu stimă,

Vasile Bratu
https://vasiledev.com
amendamax@vasiledev.com"""

    re_en_followup_subj = "Re: Automating your property data pipeline & custom Excel reporting"
    re_en_followup_body = """Hi {owner},

I wanted to quickly follow up on my previous email. Apologies for sending from a different address; I recently migrated to my official professional domain (amendamax@vasiledev.com) and wanted to make sure my message reached you.

If the team at {company_name} is still looking to automate your property data pipeline in {city}, my offer for a free 5-lead custom sample is still open.

You can also check out my latest guide on how I programmatically build these clean, client-ready Excel reporting systems:
👉 https://dev.to/amendamax2025/how-to-build-executive-ready-excel-reports-directly-in-python-using-openpyxl-393c

Let me know if you'd like me to build this free sample for you!

Best regards,

Vasile Bratu
https://vasiledev.com
amendamax@vasiledev.com"""

    # E-commerce Follow-up
    eco_it_followup_subj = "Re: Monitoraggio automatico dei prezzi concorrenti per il vostro e-commerce"
    eco_it_followup_body = """Gentile {owner},

Le scrivo per un breve follow-up al mio messaggio precedente. Mi scuso per l'invio da un indirizzo diverso; ho recentemente migrato le mie comunicazioni sulla mia email aziendale ufficiale (amendamax@vasiledev.com) e volevo assicurarmi che il mio messaggio precedente fosse arrivato.

Se il team di {company_name} desidera ancora automatizzare il monitoraggio dei prezzi dei concorrenti per proteggere i margini di profitto, la mia proposta per una demo gratuita con il tracciamento di 5 prodotti concorrenti rimane valida.

Se le interessa vedere la qualità tecnica del report Excel che genero, ho pubblicato una guida completa a riguardo qui:
👉 https://dev.to/amendamax2025/how-to-build-executive-ready-excel-reports-directly-in-python-using-openpyxl-393c

Mi faccia sapere se desidera che prepari la demo gratuita!

Cordiali saluti,

Vasile Bratu
https://vasiledev.com
amendamax@vasiledev.com"""

    eco_ro_followup_subj = "Re: Monitorizare automata preturi concurenta pentru e-commerce"
    eco_ro_followup_body = """Bună ziua {owner},

Revin cu un scurt follow-up la mesajul meu anterior. Scuze pentru trimiterea de pe o adresă diferită; tocmai am migrat comunicarea pe adresa mea profesională oficială (amendamax@vasiledev.com) și am vrut să mă asigur că mesajul meu a ajuns cu bine la dumneavoastră.

Dacă echipa {company_name} dorește în continuare să monitorizeze automat prețurile concurenței, rămâne valabilă oferta mea de a crea o monitorizare demo gratuită pentru 5 produse la alegere.

Puteți analiza calitatea tehnică a rapoartelor mele și în ghidul pe care l-am scris recent despre automatizarea Excel:
👉 https://dev.to/amendamax2025/how-to-build-executive-ready-excel-reports-directly-in-python-using-openpyxl-393c

Spuneți-mi dacă sunteți interesat de demo-ul gratuit!

Cu stimă,

Vasile Bratu
https://vasiledev.com
amendamax@vasiledev.com"""

    eco_en_followup_subj = "Re: Automating your competitor price monitoring & e-commerce reporting"
    eco_en_followup_body = """Hi {owner},

I wanted to quickly follow up on my previous email. Apologies for sending from a different address; I recently migrated to my official professional domain (amendamax@vasiledev.com) and wanted to make sure my message reached you.

If the team at {company_name} is still looking to automate competitor price monitoring to protect your margins, my offer for a free 5-product custom price tracking sample is still open.

You can also check out my latest guide on how I programmatically format these B2B executive reports in Python:
👉 https://dev.to/amendamax2025/how-to-build-executive-ready-excel-reports-directly-in-python-using-openpyxl-393c

Let me know if you'd like me to build this free sample for you!

Best regards,

Vasile Bratu
https://vasiledev.com
amendamax@vasiledev.com"""

    wa_it_subject = "Supporto tecnico Python / Web Scraping per i progetti di {company_name}"
    wa_it_body = """Gentile {owner},

Spero che questa email vi trovi bene.

Sono Vasile Bratu, uno sviluppatore Python senior specializzato in Web Scraping, automazione dati ed estrazioni complesse, residente in Italia (Cuneo).

Nel corso dei miei progetti, collaboro spesso con agenzie web e di digital marketing come {company_name} per supportare lo sviluppo tecnico di soluzioni che richiedono la raccolta automatizzata di dati, l'integrazione di API personalizzate o il superamento di barriere anti-bot complesse (Cloudflare, Akamai, Datadome).

Se il vostro team si trova a dover gestire richieste di scraping di dati, migrazioni complesse di database e-commerce, monitoraggio costante dei prezzi per conto dei vostri clienti o reporting automatizzato in Excel/Google Sheets, posso offrirvi un supporto esterno rapido e professionale per:
1. Sviluppo di scraper custom robusti (Scrapy, Playwright, Selenium).
2. Bypass di protezioni anti-bot avanzate tramite tecniche di impersonazione e proxy rotation.
3. Generazione di report Excel di livello executive direttamente via script (con libreria openpyxl).

Ho recentemente pubblicato due guide tecniche di rilievo per sviluppatori ed agenzie su Medium:
- Come bypassare Cloudflare & WAF in modo etico e sicuro:
  👉 https://medium.com/@amendamax/web-scraping-etico-e-gdpr-come-le-aziende-possono-raccogliere-dati-pubblici-online-in-totale-24715b5c76e0
- Generazione di report Excel professionali con Python:
  👉 https://medium.com/@amendamax/generazione-di-report-excel-premium-in-python-con-openpyxl-una-guida-per-sviluppatori-c44bbff633ee

Potete visionare i miei scraper open-source ed esempi di report sul mio profilo GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

La mia proposta per voi:
Offro la mia disponibilità per realizzare un piccolo test/demo gratuito di scraping su un sito target a vostra scelta (ad esempio, estrarre i primi dati da un portale che i vostri clienti monitorano), per dimostrarvi la qualità del codice e dei dati estratti senza alcun impegno da parte vostra.

Sarei felice di fare una breve chiamata conoscitiva se ritenete che possa nascere una collaborazione.

Un cordiale saluto,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub: https://github.com/amendamax"""

    wa_it_followup_subj = "Re: Supporto tecnico Python / Web Scraping per {company_name}"
    wa_it_followup_body = """Gentile {owner},

Mi permetto di fare un breve seguito alla mia email di qualche giorno fa riguardo al supporto esterno per lo sviluppo di scraper custom e automazioni dati in Python per {company_name}.

Se in questo periodo il vostro team è impegnato nello sviluppo di nuove funzionalità o ha progetti legati all'estrazione dati (ad es. per e-commerce, real estate o lead generation) in cui necessitate di una risorsa esterna specializzata per velocizzare il lavoro, sono a vostra disposizione.

In particolare, per i progetti in cui i clienti richiedono reportistica Excel avanzata, ho descritto in questo recente articolo come automatizzarla in modo elegante tramite script:
👉 https://medium.com/@amendamax/generazione-di-report-excel-premium-in-python-con-openpyxl-una-guida-per-sviluppatori-c44bbff633ee

Se desiderate effettuare la demo gratuita (un'estrazione di prova su un sito a vostra scelta per testare il bypass delle protezioni), basta che mi rispondiate indicandomi il link del sito target.

Vi ringrazio per il tempo dedicato e vi auguro una buona giornata.

Un cordiale saluto,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com"""

    wa_ro_subject = "Colaborare tehnica Python / Web Scraping pentru proiectele {company_name}"
    wa_ro_body = """Buna ziua {owner},

Numele meu este Vasile Bratu si sunt inginer software senior specializat in Web Scraping, automatizari de date si integrari de sisteme in Python.

Colaborez frecvent cu agentii web si agentii de digital marketing pentru a le ajuta sa externalizeze sarcini tehnice complexe legate de extragerea automatizata a datelor, migrarea bazelor de date, monitorizarea concurentei pentru clientii lor sau bypass-ul protectiilor anti-bot (Cloudflare, Akamai).

Va pot sustine echipa de la {company_name} ca partener tehnic extern pentru:
1. Dezvoltarea de scraper-e custom, rezistente la blocaje (Scrapy, Playwright).
2. Automatizarea rapoartelor de business direct in format Excel premium (folosind openpyxl).
3. Integrarea API-urilor si sincronizarea datelor in Cloud/CRM.

Portofoliul meu open-source si mostre de rapoarte pot fi consultate pe GitHub:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

Propunerea mea gratuita:
Sunt bucuros sa realizez un test/demo gratuit (extragerea catorva date dintr-un site target ales de dumneavoastra), pentru a va convinge de calitatea datelor livrate si a codului meu, fara nicio obligatie.

Daca doriti o scurta discutie sau un test gratuit, va rog sa imi lasati un mesaj.

Cu stima,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub: https://github.com/amendamax"""

    wa_ro_followup_subj = "Re: Colaborare tehnica Python / Web Scraping pentru {company_name}"
    wa_ro_followup_body = """Buna ziua {owner},

Revin cu un mesaj scurt in completarea emailului trimis recent cu privire la serviciile de dezvoltare software Python si web scraping pentru proiectele {company_name}.

Daca in prezent lucrati la integrari de date, monitorizare concurenti sau aveti clienti care au nevoie de extrageri masive de date si echipa interna are nevoie de suport specializat, as fi bucuros sa va ajut ca partener extern.

In special pentru proiectele care necesita rapoarte Excel executive avansate, am documentat o parte din fluxul de lucru in acest articol de pe Medium:
👉 https://medium.com/@amendamax/generazione-di-report-excel-premium-in-python-con-openpyxl-una-guida-per-sviluppatori-c44bbff633ee

Putem face oricand un test gratuit (extragerea catorva randuri dintr-un site dificil) ca sa va convingeti de fiabilitatea solutiilor mele.

Va multumesc pentru atentie si va doresc o zi excelenta.

Cu stima,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com"""

    wa_en_subject = "Python / Web Scraping technical support for {company_name} projects"
    wa_en_body = """Hi {owner},

Hope you are doing well.

I am Vasile Bratu, a senior Python developer specializing in web scraping, data pipeline automation, and anti-bot bypass (Cloudflare, Akamai).

I frequently collaborate with web development and SEO agencies like {company_name} as an external technical partner, handling complex web scraping, database migrations, competitor price monitoring for their clients, and automated Excel reporting.

I can support your team with:
1. Development of robust, production-grade custom scrapers (Scrapy, Playwright).
2. Advanced anti-bot bypass strategies (impersonation, proxy rotation).
3. Automating high-end Excel executive reports via Python scripting.

I recently published a couple of technical articles on Medium regarding GDPR compliance in scraping and professional Excel formatting:
- Web scraping ethics & GDPR:
  👉 https://medium.com/@amendamax/web-scraping-etico-e-gdpr-come-le-aziende-possono-raccogliere-dati-pubblici-online-in-totale-24715b5c76e0
- Generating premium Excel reports with Openpyxl:
  👉 https://medium.com/@amendamax/generazione-di-report-excel-premium-in-python-con-openpyxl-una-guida-per-sviluppatori-c44bbff633ee

You can review my open-source scrapers and report samples on my GitHub profile:
👉 https://github.com/amendamax/python-b2b-lead-scrapers

My free offer for you:
I'm happy to write a free proof-of-concept (POC) script to scrape a small sample from a target website of your choice, so you can evaluate the quality of the data and delivery with zero commitments.

Let me know if you would be open to a quick introductory call!

Best regards,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com
GitHub: https://github.com/amendamax"""

    wa_en_followup_subj = "Re: Python / Web Scraping technical support for {company_name}"
    wa_en_followup_body = """Hi {owner},

I'm following up on my email from a few days ago regarding external Python development and web scraping support for {company_name}.

If your team is currently handling data-heavy projects, competitor price monitoring, or requires assistance with complex anti-bot systems, I'd be happy to step in as a specialized technical partner to accelerate your deliveries.

If you have a challenging target site in mind, I can build a free extraction sample to demonstrate my capabilities. Simply reply with the target URL!

Thank you for your time, and have a great day.

Best regards,

Vasile Bratu
Senior Python & Data Automation Engineer
amendamax@vasiledev.com"""

    return {
        "outreach": {
            "Imobiliare": {
                "italy": (re_it_subject, re_it_body),
                "romania": (re_ro_subject, re_ro_body),
                "usa": (re_en_subject, re_en_body),
                "uk": (re_en_subject, re_en_body)
            },
            "E-commerce": {
                "italy": (eco_it_subject, eco_it_body),
                "romania": (eco_ro_subject, eco_ro_body),
                "usa": (eco_en_subject, eco_en_body),
                "uk": (eco_en_subject, eco_en_body)
            },
            "Web Agency": {
                "italy": (wa_it_subject, wa_it_body),
                "romania": (wa_ro_subject, wa_ro_body),
                "usa": (wa_en_subject, wa_en_body),
                "uk": (wa_en_subject, wa_en_body)
            }
        },
        "followup": {
            "Imobiliare": {
                "italy": (re_it_followup_subj, re_it_followup_body),
                "romania": (re_ro_followup_subj, re_ro_followup_body),
                "usa": (re_en_followup_subj, re_en_followup_body),
                "uk": (re_en_followup_subj, re_en_followup_body)
            },
            "E-commerce": {
                "italy": (eco_it_followup_subj, eco_it_followup_body),
                "romania": (eco_ro_followup_subj, eco_ro_followup_body),
                "usa": (eco_en_followup_subj, eco_en_followup_body),
                "uk": (eco_en_followup_subj, eco_en_followup_body)
            },
            "Web Agency": {
                "italy": (wa_it_followup_subj, wa_it_followup_body),
                "romania": (wa_ro_followup_subj, wa_ro_followup_body),
                "usa": (wa_en_followup_subj, wa_en_followup_body),
                "uk": (wa_en_followup_subj, wa_en_followup_body)
            }
        }
    }

def main():
    workspace_dir = r"C:\Users\bratu\Documents\antigravity\amazing-borg"
    
    # Gather all leads dynamically using glob
    import glob
    leads = []
    excel_files = glob.glob(os.path.join(workspace_dir, "leads_*.xlsx"))
    outreach_file = os.path.join(workspace_dir, "PROSPECTE_AGENTII_OUTREACH.xlsx")
    if os.path.exists(outreach_file):
        excel_files.append(outreach_file)
    for file_path in excel_files:
        filename = os.path.basename(file_path).lower()
        if "ecommerce" in filename:
            niche_type = "E-commerce"
        elif "webagency" in filename or "agency" in filename:
            niche_type = "Web Agency"
        else:
            niche_type = "Imobiliare"
        leads.extend(get_leads_from_excel(file_path, niche_type))
    
    if not leads:
        print("No leads found in Excel sheets!")
        return
        
    templates = get_templates()
    
    # Generate HTML content
    html_content = """<!DOCTYPE html>
<html lang="ro">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Dashboard Lansare Outreach B2B - Vasile Bratu</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #0f172a;
            color: #f8fafc;
            margin: 0;
            padding: 20px;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
        }
        header {
            text-align: center;
            padding: 30px 0;
            border-bottom: 1px solid #1e293b;
        }
        h1 {
            color: #38bdf8;
            margin-bottom: 5px;
            font-size: 2.5em;
        }
        p.subtitle {
            color: #94a3b8;
            font-size: 1.1em;
            margin-top: 5px;
        }
        .stats {
            display: flex;
            justify-content: space-around;
            margin: 30px 0;
            background: rgba(30, 41, 59, 0.5);
            padding: 15px;
            border-radius: 12px;
            border: 1px solid #334155;
        }
        .stat-item {
            text-align: center;
        }
        .stat-val {
            font-size: 1.8em;
            font-weight: bold;
            color: #34d399;
        }
        .stat-lbl {
            color: #94a3b8;
            font-size: 0.9em;
        }
        .grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(360px, 1fr));
            gap: 20px;
            margin-top: 30px;
        }
        .card {
            background: #1e293b;
            border: 1px solid #334155;
            border-radius: 12px;
            padding: 20px;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            transition: transform 0.2s, box-shadow 0.2s;
        }
        .card:hover {
            transform: translateY(-5px);
            box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.3);
            border-color: #38bdf8;
        }
        .card-header {
            border-bottom: 1px solid #334155;
            padding-bottom: 10px;
            margin-bottom: 15px;
        }
        .comp-name {
            font-size: 1.2em;
            font-weight: bold;
            color: #f1f5f9;
        }
        .badge {
            display: inline-block;
            padding: 3px 8px;
            border-radius: 20px;
            font-size: 0.8em;
            font-weight: bold;
            margin-top: 5px;
        }
        .badge-it { background-color: #0d9488; color: #f0fdfa; }
        .badge-ro { background-color: #dc2626; color: #fef2f2; }
        .badge-re { background-color: #059669; color: #ecfdf5; }
        .badge-eco { background-color: #d97706; color: #fffbeb; }
        .badge-wa { background-color: #6366f1; color: #e0e7ff; }
        .badge-uk { background-color: #1e3a8a; color: #dbeafe; }
        .badge-usa { background-color: #3b82f6; color: #eff6ff; }
        .details {
            font-size: 0.95em;
            color: #cbd5e1;
            line-height: 1.6;
        }
        .details strong {
            color: #94a3b8;
        }
        .btn {
            display: block;
            text-align: center;
            background-color: #0284c7;
            color: white;
            text-decoration: none;
            padding: 12px;
            border-radius: 8px;
            font-weight: bold;
            margin-top: 10px;
            transition: background-color 0.2s;
        }
        .btn:hover {
            background-color: #0369a1;
        }
        .btn-send {
            background-color: #10b981;
            margin-top: 15px;
        }
        .btn-send:hover {
            background-color: #059669;
        }
        .btn-followup {
            background-color: #d97706;
        }
        .btn-followup:hover {
            background-color: #b45309;
        }
        /* State sent styles */
        .card.state-sent .btn-send {
            background-color: #065f46 !important;
            color: #a7f3d0;
        }
        /* State followup styles */
        .card.state-followup .btn-followup {
            background-color: #065f46 !important;
            color: #a7f3d0;
        }
        
        /* Tabs Styling */
        .tabs {
            display: flex;
            justify-content: center;
            margin-bottom: 30px;
            gap: 15px;
        }
        .tab-btn {
            background-color: #1e293b;
            border: 1px solid #334155;
            color: #94a3b8;
            padding: 12px 24px;
            font-size: 1.1em;
            font-weight: bold;
            border-radius: 8px;
            cursor: pointer;
            transition: all 0.2s ease;
        }
        .tab-btn:hover {
            border-color: #38bdf8;
            color: #f1f5f9;
        }
        .tab-btn.active {
            background-color: #0284c7;
            border-color: #38bdf8;
            color: white;
            box-shadow: 0 4px 6px -1px rgba(2, 132, 199, 0.4);
        }
        .tab-content {
            display: none;
        }
        .tab-content.active-content {
            display: block;
        }
        
        /* Card Actions & Sent State */
        .card-actions {
            display: flex;
            flex-direction: column;
            gap: 10px;
            margin-top: 15px;
        }
        .btn-sent-toggle {
            background-color: #334155;
            color: #cbd5e1;
            border: 1px solid #475569;
            padding: 8px;
            border-radius: 6px;
            font-size: 0.9em;
            font-weight: bold;
            cursor: pointer;
            transition: all 0.2s;
        }
        .btn-sent-toggle:hover {
            background-color: #475569;
            color: white;
        }
        .btn-sent-toggle.btn-sent-active {
            background-color: #065f46;
            border-color: #059669;
            color: #a7f3d0;
        }
        .card.card-sent {
            opacity: 0.45;
            border-color: #1e293b;
            filter: grayscale(40%);
            transition: all 0.3s ease;
        }
        .card.card-sent:hover {
            opacity: 0.85;
            filter: none;
        }
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>✉️ Lansator Outreach B2B Global</h1>
            <p class="subtitle">Selectează o firmă calificată de mai jos pentru a deschide e-mailul pre-formatat cu portofoliul tău</p>
        </header>
        
        <div class="stats">
            <div class="stat-item">
                <div class="stat-val">""" + str(len(leads)) + """</div>
                <div class="stat-lbl">Total Prospecte Calificate</div>
            </div>
            <div class="stat-item">
                <div class="stat-val">""" + str(len([l for l in leads if l["country"] == "italy"])) + """</div>
                <div class="stat-lbl">Italia (IT)</div>
            </div>
            <div class="stat-item">
                <div class="stat-val">""" + str(len([l for l in leads if l["country"] == "romania"])) + """</div>
                <div class="stat-lbl">România (RO)</div>
            </div>
            <div class="stat-item">
                <div class="stat-val">""" + str(len([l for l in leads if l["country"] in ["usa", "uk"]])) + """</div>
                <div class="stat-lbl">Global (EN)</div>
            </div>
            <div class="stat-item">
                <div class="stat-val">""" + str(len([l for l in leads if l["niche"] == "Imobiliare"])) + """</div>
                <div class="stat-lbl">Imobiliare (RE)</div>
            </div>
            <div class="stat-item">
                <div class="stat-val">""" + str(len([l for l in leads if l["niche"] == "E-commerce"])) + """</div>
                <div class="stat-lbl">E-commerce (ECO)</div>
            </div>
        </div>
        
        <!-- Tab Navigation Buttons -->
        <div class="tabs">
            <button class="tab-btn active" onclick="openTab(event, 'italy-tab')">Italia 🇮🇹</button>
            <button class="tab-btn" onclick="openTab(event, 'romania-tab')">România 🇷🇴</button>
            <button class="tab-btn" onclick="openTab(event, 'global-tab')">Internațional 🌐</button>
        </div>
"""
    
    # Partition leads by group
    italy_leads = [l for l in leads if l["country"] == "italy"]
    romania_leads = [l for l in leads if l["country"] == "romania"]
    global_leads = [l for l in leads if l["country"] in ["usa", "uk"]]
    
    def render_leads_group(group_leads):
        group_html = ""
        for l in group_leads:
            subj, body = templates["outreach"][l["niche"]][l["country"]]
            followup_subj, followup_body = templates["followup"][l["niche"]][l["country"]]
            
            # Handle empty/None owner gracefully
            owner_val = l["owner"]
            if not owner_val or owner_val == "None" or str(owner_val).strip() == "":
                if l["country"] == "italy":
                    if l["niche"] == "Web Agency":
                        owner_val = "Responsabile Tecnico"
                    else:
                        owner_val = "Responsabile"
                elif l["country"] == "romania":
                    if l["niche"] == "Web Agency":
                        owner_val = "Responsabil Tehnic"
                    else:
                        owner_val = f"Echipa {l['company_name']}"
                else:
                    if l["niche"] == "Web Agency":
                        owner_val = "Technical Lead"
                    else:
                        owner_val = f"Team at {l['company_name']}"
            
            # Populate template parameters (Outreach)
            formatted_subj = subj
            formatted_body = body.format(
                owner=owner_val,
                city=l["city"],
                company_name=l["company_name"]
            )
            
            # URL encode subject and body for mailto link
            encoded_subj = urllib.parse.quote(formatted_subj)
            encoded_body = urllib.parse.quote(formatted_body)
            mail_link = f"mailto:{l['email']}?subject={encoded_subj}&body={encoded_body}"

            # Populate template parameters (Followup)
            formatted_followup_subj = followup_subj
            formatted_followup_body = followup_body.format(
                owner=owner_val,
                city=l["city"],
                company_name=l["company_name"]
            )
            encoded_followup_subj = urllib.parse.quote(formatted_followup_subj)
            encoded_followup_body = urllib.parse.quote(formatted_followup_body)
            followup_mail_link = f"mailto:{l['email']}?subject={encoded_followup_subj}&body={encoded_followup_body}"
            
            if l["country"] == "italy":
                badge_class = "badge-it"
                country_lbl = "Italia 🇮🇹"
            elif l["country"] == "romania":
                badge_class = "badge-ro"
                country_lbl = "România 🇷🇴"
            elif l["country"] == "uk":
                badge_class = "badge-uk"
                country_lbl = "Marea Britanie 🇬🇧"
            else: # usa
                badge_class = "badge-usa"
                country_lbl = "SUA 🇺🇸"
            niche_badge_class = "badge-wa" if l["niche"] == "Web Agency" else "badge-re" if l["niche"] == "Imobiliare" else "badge-eco"
            
            group_html += f"""
                <div class="card" data-company="{l['company_name']}">
                    <div>
                        <div class="card-header">
                            <div class="comp-name">{l['company_name']}</div>
                            <span class="badge {badge_class}">{country_lbl} - {l['city']}</span>
                            <span class="badge {niche_badge_class}">{l['niche']}</span>
                        </div>
                        <div class="details">
                            <strong>Proprietar / Manager:</strong> {l['owner'] if l['owner'] and l['owner'] != 'None' else 'Echipă / Nespecificat'}<br>
                            <strong>Email:</strong> {l['email']}<br>
                            <strong>Telefon:</strong> {l['phone']}<br>
                            <strong>Site:</strong> <a href="{l['website']}" target="_blank" style="color: #38bdf8; text-decoration: none;">{l['clean_domain'] if 'clean_domain' in l else l['website']} ↗</a>
                        </div>
                    </div>
                    <div class="card-actions">
                        <a href="{mail_link}" class="btn btn-send" onclick="markAutoSent('{l['company_name']}')">Trimite Email ✉️</a>
                        <a href="{followup_mail_link}" class="btn btn-followup" onclick="markAutoFollowup('{l['company_name']}')">Trimite Follow-up 🔁</a>
                        <button class="btn-sent-toggle">Marchează ca Trimis</button>
                    </div>
                </div>
            """
        return group_html

    # Render Italy Tab
    html_content += """
        <div id="italy-tab" class="tab-content active-content">
            <div class="grid">
    """ + render_leads_group(italy_leads) + """
            </div>
        </div>
    """
    
    # Render Romania Tab
    html_content += """
        <div id="romania-tab" class="tab-content">
            <div class="grid">
    """ + render_leads_group(romania_leads) + """
            </div>
        </div>
    """
    
    # Render Global Tab
    html_content += """
        <div id="global-tab" class="tab-content">
            <div class="grid">
    """ + render_leads_group(global_leads) + """
            </div>
        </div>
    """
        
    html_content += """
    </div>
    
    <script>
        // Tab switching logic
        function openTab(evt, tabId) {
            // Hide all tab contents
            const contents = document.querySelectorAll('.tab-content');
            contents.forEach(c => {
                c.classList.remove('active-content');
            });
            
            // Deactivate all tab buttons
            const buttons = document.querySelectorAll('.tab-btn');
            buttons.forEach(b => {
                b.classList.remove('active');
            });
            
            // Show selected tab content and active button
            document.getElementById(tabId).classList.add('active-content');
            evt.currentTarget.classList.add('active');
        }
        
        // Auto mark as sent on click (optional but nice)
        function markAutoSent(companyName) {
            const card = document.querySelector(`[data-company="${companyName}"]`);
            if (card) {
                const sentBtn = card.querySelector('.btn-sent-toggle');
                card.classList.add('card-sent');
                card.classList.add('state-sent');
                card.classList.remove('state-followup');
                if (sentBtn) {
                    sentBtn.classList.add('btn-sent-active');
                    sentBtn.innerHTML = 'Trimis ✔️';
                }
                localStorage.setItem('sent_' + companyName, 'true');
            }
        }

        // Auto mark as follow-up sent
        function markAutoFollowup(companyName) {
            const card = document.querySelector(`[data-company="${companyName}"]`);
            if (card) {
                const sentBtn = card.querySelector('.btn-sent-toggle');
                card.classList.add('card-sent');
                card.classList.remove('state-sent');
                card.classList.add('state-followup');
                if (sentBtn) {
                    sentBtn.classList.add('btn-sent-active');
                    sentBtn.innerHTML = 'Follow-up Trimis ✔️';
                }
                localStorage.setItem('sent_' + companyName, 'followup');
            }
        }
        
        // Persistent Sent Tracking using localStorage
        document.addEventListener("DOMContentLoaded", function() {
            // Auto-deselect recently repaired emails
            const repaired = [
                "Savedbythedress",
                "Flamingoshoptorino",
                "The Realestateagency",
                "Top10Propertyagents",
                "Chestertons",
                "Floridarealtyofmiami",
                "Vestique",
                "Miss Rosier",
                "Erdem"
            ];
            repaired.forEach(comp => {
                localStorage.removeItem("sent_" + comp);
            });

            const cards = document.querySelectorAll(".card");
            
            cards.forEach(card => {
                const companyName = card.getAttribute("data-company");
                const sentBtn = card.querySelector('.btn-sent-toggle');
                
                // Check if already sent in localStorage
                const sentState = localStorage.getItem('sent_' + companyName);
                if (sentState === 'true') {
                    card.classList.add('card-sent');
                    card.classList.add('state-sent');
                    if (sentBtn) {
                        sentBtn.classList.add('btn-sent-active');
                        sentBtn.innerHTML = 'Trimis ✔️';
                    }
                } else if (sentState === 'followup') {
                    card.classList.add('card-sent');
                    card.classList.add('state-followup');
                    if (sentBtn) {
                        sentBtn.classList.add('btn-sent-active');
                        sentBtn.innerHTML = 'Follow-up Trimis ✔️';
                    }
                }
                
                if (sentBtn) {
                    sentBtn.addEventListener('click', function(e) {
                        e.preventDefault();
                        e.stopPropagation();
                        
                        const isSent = card.classList.toggle('card-sent');
                        sentBtn.classList.toggle('btn-sent-active');
                        
                        if (isSent) {
                            // Cycle state: unsent -> sent -> followup -> unsent
                            const currentState = localStorage.getItem('sent_' + companyName);
                            if (!currentState) {
                                sentBtn.innerHTML = 'Trimis ✔️';
                                card.classList.add('state-sent');
                                card.classList.remove('state-followup');
                                localStorage.setItem('sent_' + companyName, 'true');
                            } else if (currentState === 'true') {
                                sentBtn.innerHTML = 'Follow-up Trimis ✔️';
                                card.classList.remove('state-sent');
                                card.classList.add('state-followup');
                                localStorage.setItem('sent_' + companyName, 'followup');
                            } else {
                                card.classList.remove('card-sent');
                                card.classList.remove('state-sent');
                                card.classList.remove('state-followup');
                                sentBtn.classList.remove('btn-sent-active');
                                sentBtn.innerHTML = 'Marchează ca Trimis';
                                localStorage.removeItem('sent_' + companyName);
                            }
                        } else {
                            card.classList.remove('state-sent');
                            card.classList.remove('state-followup');
                            sentBtn.innerHTML = 'Marchează ca Trimis';
                            localStorage.removeItem('sent_' + companyName);
                        }
                    });
                }
            });
        });
    </script>
</body>
</html>
"""
    
    # Save the HTML dashboard to Desktop/Outreach_B2B
    outreach_dir = os.path.join(os.path.expanduser("~"), "Desktop", "Outreach_B2B")
    os.makedirs(outreach_dir, exist_ok=True)
    desktop_html = os.path.join(outreach_dir, "TRIMITE_EMAILURI_B2B.html")
    with open(desktop_html, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"Interactive outreach dashboard generated successfully at: {desktop_html}")

if __name__ == "__main__":
    main()
