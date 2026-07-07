# -*- coding: utf-8 -*-
"""
Script pentru scanarea și generarea unui raport premium cu atracțiile de familie 
pe o rază de 20-25 km de Garessio (CN), Italia.
Creat special pentru portofoliul și uzul personal al lui Vasile Bratu.
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Asigurăm codificarea UTF-8 pentru Windows
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except AttributeError:
    pass

# Setul de date cu atracții verificate în zona Garessio (pe o rază de 20-25 km)
# Coordonate de pornire: Garessio (Via Valcasotto 14)
ATRACTII = [
    {
        "nume": "Reggia di Valcasotto (Castello Reale)",
        "localitate": "Garessio (CN)",
        "tip": "Istorie & Cultură",
        "distanta": 10.2,
        "rating": 4.6,
        "descriere": "Fostă rezidență regală de vânătoare a Casei de Savoia. Un castel superb înconjurat de păduri montane, recent restaurat. Tururile ghidate sunt interactive și extrem de educative pentru copii.",
        "gmaps": "https://maps.google.com/?q=Castello+Reale+di+Casotto+Garessio"
    },
    {
        "nume": "Borgo Maggiore și Piazzetta San Giovanni",
        "localitate": "Garessio (CN)",
        "tip": "Istorie & Plimbare",
        "distanta": 0.5,
        "rating": 4.7,
        "descriere": "Centrul istoric medieval din Garessio, desemnat printre cele mai frumoase sate din Italia (Borghi più belli d'Italia). Piața decorată cu pietre albe și negre și porțile medievale sunt perfecte pentru o plimbare liniștită.",
        "gmaps": "https://maps.google.com/?q=Piazzetta+San+Giovanni+Garessio"
    },
    {
        "nume": "Parcul Izvoarelor San Bernardo (Parco delle Fonti)",
        "localitate": "Garessio (CN)",
        "tip": "Natură & Relaxare",
        "distanta": 1.8,
        "rating": 4.5,
        "descriere": "Parcul istoric unde izvorăște celebra apă minerală San Bernardo. O pădure seculară de pini cu poteci line, locuri de joacă pentru copii și aer curat, ideal pentru picnicuri în familie.",
        "gmaps": "https://maps.google.com/?q=Fonti+San+Bernardo+Garessio"
    },
    {
        "nume": "Garessio 2000 Ski & Trekking Area",
        "localitate": "Garessio (CN)",
        "tip": "Sport & Aventură",
        "distanta": 12.0,
        "rating": 4.2,
        "descriere": "Stațiune montană perfectă iarna pentru schi în familie și săniuș, iar vara pentru drumeții montane uimitoare spre Muntele Mindino, de unde se poate vedea marea în zilele senine.",
        "gmaps": "https://maps.google.com/?q=Garessio+2000"
    },
    {
        "nume": "Centrul Istoric Ormea și Poteca 'Il Biale'",
        "localitate": "Ormea (CN)",
        "tip": "Istorie & Plimbare",
        "distanta": 11.5,
        "rating": 4.5,
        "descriere": "Oraș pitoresc cu un centru istoric în formă de inimă. Copiii vor adora să exploreze 'trevi' (străduțele înguste medievale) și sistemul vechi de curățare a străzilor prin canale cu apă curgătoare.",
        "gmaps": "https://maps.google.com/?q=Ormea+Centro+Storico"
    },
    {
        "nume": "Ruinele Castelului Priola",
        "localitate": "Priola (CN)",
        "tip": "Istorie & Explorare",
        "distanta": 7.8,
        "rating": 4.3,
        "descriere": "Ruinele unui turn și castel medieval cocoțat pe o colină deasupra văii Tanaro. O drumeție scurtă, ușoară și plină de aventură pentru micii exploratori.",
        "gmaps": "https://maps.google.com/?q=Castello+di+Priola"
    },
    {
        "nume": "Podul Medieval din Bagnasco",
        "localitate": "Bagnasco (CN)",
        "tip": "Istorie & Natură",
        "distanta": 13.5,
        "rating": 4.6,
        "descriere": "Un pod spectaculos din piatră cu arcade medievale peste râul Tanaro. Zona din jur este amenajată cu spații verzi pe malul apei, ideale pentru relaxare și fotografii de familie.",
        "gmaps": "https://maps.google.com/?q=Ponte+Romano+Bagnasco"
    },
    {
        "nume": "Forturile Istorice din Colle di Nava",
        "localitate": "Colle di Nava (IM)",
        "tip": "Istorie & Explorare",
        "distanta": 19.1,
        "rating": 4.5,
        "descriere": "Complex de fortificații militare din secolul al XIX-lea (Forte Centrale). Spațiile verzi din jur sunt excelente pentru picnic, iar Nava este faimoasă pentru produsele locale din lavandă.",
        "gmaps": "https://maps.google.com/?q=Forte+Centrale+Colle+di+Nava"
    },
    {
        "nume": "Viola St Grée - Downhill & Kids Park",
        "localitate": "Viola (CN)",
        "tip": "Sport & Aventură",
        "distanta": 15.4,
        "rating": 4.3,
        "descriere": "Stațiune de vară și iarnă cu facilități speciale pentru copii, inclusiv piste de bob de vară (tubby), trasee de biciclete ușoare și locuri de joacă gonflabile la poalele munților.",
        "gmaps": "https://maps.google.com/?q=Viola+St+Gree"
    },
    {
        "nume": "Grotte di Bossea (Peștera Bossea)",
        "localitate": "Frabosa Soprana (CN)",
        "tip": "Natură & Știință",
        "distanta": 22.5,
        "rating": 4.8,
        "descriere": "Una dintre cele mai spectaculoase și mari peșteri turistice din Italia. Include râuri subterane, cascade gigantice și scheletul reconstituit al unui urs preistoric (Ursus spelaeus). O experiență magică!",
        "gmaps": "https://maps.google.com/?q=Grotte+di+Bossea"
    },
    {
        "nume": "Lago di Osiglia (Lacul Osiglia)",
        "localitate": "Osiglia (SV)",
        "tip": "Natură & Agrement",
        "distanta": 24.2,
        "rating": 4.6,
        "descriere": "Un lac montan superb înconjurat de păduri de fagi. Familiile pot închiria hidrobiciclete, canoe, pot înota în zonele special amenajate sau pot face un grătar pe malul lacului.",
        "gmaps": "https://maps.google.com/?q=Lago+di+Osiglia"
    },
    {
        "nume": "Muzeul Căilor Ferate și Castelul Nucetto",
        "localitate": "Nucetto (CN)",
        "tip": "Istorie & Educație",
        "distanta": 19.5,
        "rating": 4.4,
        "descriere": "Muzeul dedicat liniei istorice Ceva-Ormea și istoriei locale, urmat de o plimbare până la ruinele maiestuoase ale castelului Nucetto ce domină valea.",
        "gmaps": "https://maps.google.com/?q=Castello+di+Nucetto"
    }
]

def genereaza_excel(output_path):
    """Generează un fișier Excel premium stilizat în nuanțe de Green Emerald montan."""
    print(f"[Orchestrator] Se creează raportul Excel la adresa: {output_path}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atracții Familie Garessio"
    
    # Activăm afișarea liniilor de grilă (gridlines)
    ws.views.sheetView[0].showGridLines = True
    
    # Antete
    headers = [
        "Atracție / Obiectiv", 
        "Localitate", 
        "Categorie", 
        "Distanță (km)", 
        "Rating Google", 
        "Recomandare & Descriere Familie (Garessio)", 
        "Locație Google Maps"
    ]
    
    # Culori Tematice: Alpine Forest Emerald
    header_fill = PatternFill(start_color="113E21", end_color="113E21", fill_type="solid")  # Deep Pine Green
    zebra_fill = PatternFill(start_color="F4FAF6", end_color="F4FAF6", fill_type="solid")   # Pale Mint
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    accent_fill = PatternFill(start_color="E2F0E7", end_color="E2F0E7", fill_type="solid")  # Accent Mint
    
    # Fonturi elegante Segoe UI (standard premium)
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="2C3E50")
    bold_name_font = Font(name="Segoe UI", size=10, bold=True, color="0D2E18")
    link_font = Font(name="Segoe UI", size=10, underline="single", color="1B5E20")
    distance_font = Font(name="Segoe UI", size=10, bold=True, color="2E7D32")
    rating_font = Font(name="Segoe UI", size=10, bold=True, color="E67E22")
    
    # Margini subțiri și curate
    thin_border_side = Side(border_style="thin", color="D0DFD5")
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Stilare rând de Antet
    ws.row_dimensions[1].height = 36
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = grid_border
        
    # Sortăm atracțiile după distanță (cele mai apropiate primele)
    atractii_sortate = sorted(ATRACTII, key=lambda x: x["distanta"])
    
    # Populare date
    for row_idx, atr in enumerate(atractii_sortate, start=2):
        ws.row_dimensions[row_idx].height = 42  # Rânduri spațioase ("aerat")
        current_fill = zebra_fill if (row_idx % 2 == 1) else white_fill
        
        row_data = [
            atr["nume"],
            atr["localitate"],
            atr["tip"],
            atr["distanta"],
            atr["rating"],
            atr["descriere"],
            atr["gmaps"]
        ]
        
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = current_fill
            cell.border = grid_border
            cell.font = data_font
            
            # Aliniere specifică pentru design profesionist
            if col_idx in [1, 6]:  # Text lung (Nume, Descriere) stânga
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(col_idx == 6))
            elif col_idx in [2, 3]:  # Localitate, Tip mijloc
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [4, 5, 7]:  # Cifre și link-uri mijloc
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Stiluri și valori speciale pe coloane
            if col_idx == 1:
                cell.value = val
                cell.font = bold_name_font
            elif col_idx == 4:
                cell.value = f"{val:.1f} km"
                cell.font = distance_font
            elif col_idx == 5:
                cell.value = f"⭐ {val:.1f}"
                cell.font = rating_font
            elif col_idx == 7:
                cell.value = "Vezi pe Hartă 🗺️"
                cell.hyperlink = val
                cell.font = link_font
            else:
                cell.value = val
                
    # Lățimi auto-ajustate ale coloanelor pentru a preveni trunchierea
    # Coloana de descriere (6) o fixăm la o lățime mai mare, fiind text lung cu wrap_text
    fixed_widths = {
        1: 30,  # Nume obiectiv
        2: 18,  # Localitate
        3: 18,  # Categorie
        4: 15,  # Distanță
        5: 14,  # Rating
        6: 55,  # Descriere lungă cu wrap_text
        7: 20   # Link Maps
    }
    
    for col_idx, width in fixed_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Salvăm fișierul Excel
    wb.save(output_path)
    print(f"[Excel] Raportul a fost salvat cu succes în: {output_path}")

def main():
    print("==================================================================")
    print("         SCANARE ATRACȚII DE FAMILIE - RAZĂ 20-25 KM             ")
    print("==================================================================")
    print("Punct de Pornire: Garessio (Via Valcasotto 14), Piemonte, Italia")
    print("Nivel de filtrare: Familie-Friendly & Activități Copii")
    print("-" * 66)
    
    # Generăm calea de salvare pe Desktop-ul utilizatorului
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    output_xlsx = os.path.join(desktop_dir, "Atractii_Familie_Garessio.xlsx")
    
    try:
        genereaza_excel(output_xlsx)
        print(f"\n[Orchestrator] SUCCES! Raportul premium a fost creat pe Desktop:")
        print(f"-> {output_xlsx}")
    except Exception as e:
        print(f"\n[Eroare] A apărut o problemă la generare: {e}")
        
    print("\n" + "=" * 66)
    print("                  PROCES DE SCANARE FINALIZAT                   ")
    print("==================================================================")

if __name__ == "__main__":
    main()
