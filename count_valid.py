import glob
import openpyxl
import os

def main():
    files = glob.glob("leads_*.xlsx")
    total_valid = 0
    ignore_domains = ['example.com', 'domain.com', 'yourcompany.co.uk', 'yourdomain.com', 'yourdomain', 'example']
    
    for f in files:
        if "contacted_success" in f:
            continue
        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            file_valid = 0
            
            # Find email column index
            email_col = None
            for col in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=col).value).lower()
                if "email" in header or "e-mail" in header:
                    email_col = col
                    break
                    
            if not email_col:
                # fallback: search first row data
                for col in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=2, column=col).value)
                    if "@" in val:
                        email_col = col
                        break
                        
            if email_col:
                for row in range(2, ws.max_row + 1):
                    email = str(ws.cell(row=row, column=email_col).value).strip()
                    if email and email != "N/A" and email != "None" and "@" in email:
                        # check if it's a placeholder
                        is_placeholder = any(ig in email.lower() for ig in ignore_domains)
                        if not is_placeholder:
                            file_valid += 1
            total_valid += file_valid
            print(f"{os.path.basename(f)}: {file_valid} valid emails")
        except Exception as e:
            print(f"Error reading {f}: {e}")
            
    print(f"\nTotal valid, non-placeholder emails across all files: {total_valid}")

if __name__ == "__main__":
    main()
