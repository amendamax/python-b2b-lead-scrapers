import os
import json
import urllib.parse
import openpyxl
import re

def clean_js_string(text):
    if not text:
        return ""
    # Escape single quotes and backslashes for JS strings
    return text.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n").replace("\r", "")

def clean_decident_name(name):
    if not name or name == "N/A":
        return "N/A"
    stop_words = {
        "read", "bio", "view", "profile", "click", "more", "learn", "about", 
        "contact", "team", "at", "crif", "svolgimento", "non", "raccoglie", 
        "newly", "controlling", "is", "operates", "allo", "di", "quindi", 
        "bar", "iqo", "am", "are", "the", "and", "for", "to", "or", "in", "page"
    }
    words = re.sub(r'[^\w\s]', ' ', name).split()
    cleaned_parts = [w for w in words if w.lower() not in stop_words]
    final_parts = [p for p in cleaned_parts if p[0].isupper() and len(p) >= 2]
    return " ".join(final_parts) if final_parts else "N/A"

def get_email_templates(company_name, decident, language):
    clean_name = clean_decident_name(decident)
    salutation = clean_name.split()[0] if (clean_name and clean_name != "N/A") else ""
    
    if language == "italian":
        salute = f"Ciao {salutation}" if salutation else "Gentile Team de"
        subject = f"Collaborazione Web Scraping & Automatizzazione Dati per {company_name}"
        body = f"""{salute} {company_name},

Vi contatto perché apprezzo molto i progetti digitali che realizzate.

Spesso le agenzie web e di marketing si scontrano con problemi legati al blocco dei flussi di dati da parte di sistemi anti-bot (come Cloudflare o Akamai) o perdono ore preziose a formattare manualmente report in Excel per i clienti.

Mi chiamo Vasile e sono uno **Specialista in Python Web Scraping e Automatizzazione Dati**. Aiuto le agenzie a sviluppare motori di scraping personalizzati e resilienti che superano i moderni sistemi di sicurezza (utilizzando Playwright stealth, configurazioni proxy e manipolazione delle impronte digitali TLS) e che generano report pronti per i clienti in formati premium.

Potete visionare alcuni dei miei approfondimenti tecnici direttamente sul mio portfolio: https://vasiledev.com o sul mio profilo LinkedIn.

Avete attualmente progetti attivi o clienti che necessitano di estrazione dati su misura o di dashboard automatizzate? Sarei felice di collaborare con voi.

Un cordiale saluto,

Vasile Bratu
Sviluppatore Python & Data Automation
amendamax@vasiledev.com
"""
    elif language == "romanian":
        salute = f"Salut {salutation}" if salutation else "Bună ziua, echipa"
        subject = f"Colaborare Web Scraping și Automatizare Date pentru {company_name}"
        body = f"""{salute} {company_name},

Vă contactez pentru că apreciez portofoliul de proiecte digitale pe care le realizați.

Multe agenții web și de marketing se confruntă des cu probleme de blocare a scripturilor de colectare a datelor de către sisteme anti-bot (precum Cloudflare sau Akamai) sau pierd timp prețios formatând manual rapoarte în Excel pentru clienți.

Numele meu este Vasile și sunt **Specialist în Python Web Scraping și Automatizare**. Ajut agențiile să construiască motoare de scraping personalizate și rezistente, care ocolesc protecțiile moderne (folosind Playwright stealth, proxy-uri rezidențiale și TLS fingerprinting) și livrează datele direct în rapoarte premium, complet formatate.

Puteți vedea câteva dintre analizele mele tehnice direct pe site-ul meu: https://vasiledev.com sau pe profilul meu de LinkedIn.

Aveți în acest moment proiecte active sau clienți care au nevoie de extragere de date personalizată sau de rapoarte automatizate? Mi-ar plăcea să colaborăm.

O zi excelentă,

Vasile Bratu
Python & Data Automation Engineer
amendamax@vasiledev.com
"""
    else: # English default
        salute = f"Hi {salutation}" if salutation else "Hi Team"
        subject = f"Custom Python Scrapers & Anti-Bot Bypass for {company_name}"
        body = f"""{salute},

I came across {company_name} and love the digital work you do.

Many web and marketing agencies often face issues with data collection pipelines getting blocked by Cloudflare/Akamai, or they waste hours manually formatting Excel dashboards for their clients.

I’m Vasile, a **Python Web Scraping & Data Automation Specialist**. I help agencies build resilient, custom scraping engines that bypass modern anti-bots (using Playwright, stealth setups, and custom TLS fingerprinting) and output clean, executive-ready reports.

You can check out some of my technical breakdowns on my website: https://vasiledev.com or on my LinkedIn profile.

Do you have any active projects or clients requiring custom data extraction or automated reports right now? I would love to help.

Best regards,

Vasile Bratu
Python & Data Automation Specialist
amendamax@vasiledev.com
"""
    return subject, body

def main():
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    excel_path = os.path.join(desktop_dir, "PROSPECTE_AGENTII_OUTREACH.xlsx")
    
    if not os.path.exists(excel_path):
        print(f"[!] Fișierul Excel nu a fost găsit pe Desktop: {excel_path}")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    leads = []
    
    for row in range(2, ws.max_row + 1):
        company = ws.cell(row=row, column=1).value
        location = ws.cell(row=row, column=2).value
        website = ws.cell(row=row, column=3).value
        email_addr = ws.cell(row=row, column=4).value
        if email_addr:
            email_addr_low = str(email_addr).lower().strip()
            placeholders = ["exemplu.ro", "example.com", "yourdomain.com", "domain.com", "email.com", "numecomerciant", "numesite", "numedomeniu", "email@email"]
            if any(p in email_addr_low for p in placeholders) or email_addr_low == "n/a":
                email_addr = "N/A"
        phone = ws.cell(row=row, column=5).value
        decident = ws.cell(row=row, column=6).value
        linkedin = ws.cell(row=row, column=7).value
        source_url = ws.cell(row=row, column=8).value
        
        if not company:
            continue
            
        location_low = str(location).lower()
        if "italy" in location_low or "milano" in location_low:
            lang = "italian"
            city_filter = "Milano"
        elif "romania" in location_low or "bucurești" in location_low:
            lang = "romanian"
            city_filter = "București"
        elif "london" in location_low:
            lang = "english"
            city_filter = "London"
        else:
            lang = "english"
            city_filter = "New York"
            
        subject, body = get_email_templates(company, decident, lang)
        
        leads.append({
            "company": company,
            "location": location,
            "city_filter": city_filter,
            "website": website,
            "email": email_addr if email_addr else "N/A",
            "phone": phone if phone else "N/A",
            "decident": decident if decident else "N/A",
            "linkedin": linkedin if linkedin else "N/A",
            "source_url": source_url if source_url else website,
            "subject": subject,
            "body": body
        })

    # Generate HTML content
    html_template = """<!DOCTYPE html>
<html lang="ro">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Outreach B2B Dashboard - Vasile Bratu</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;800&display=swap" rel="stylesheet">
    <style>
        :root {
            --primary: #1B4D3E;
            --primary-light: #2c725c;
            --accent: #E2F0EA;
            --bg-dark: #091310;
            --card-bg: rgba(27, 77, 62, 0.1);
            --card-border: rgba(27, 77, 62, 0.25);
            --text-main: #E2F0EA;
            --text-muted: #8AAEA1;
            --transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }

        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
            font-family: 'Outfit', sans-serif;
        }

        body {
            background-color: var(--bg-dark);
            color: var(--text-main);
            min-height: 100vh;
            padding: 2rem;
            background-image: radial-gradient(circle at 10% 20%, rgba(27, 77, 62, 0.15) 0%, transparent 40%),
                              radial-gradient(circle at 90% 80%, rgba(27, 77, 62, 0.15) 0%, transparent 40%);
            background-attachment: fixed;
        }

        header {
            max-width: 1200px;
            margin: 0 auto 3rem auto;
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 1px solid var(--card-border);
            padding-bottom: 1.5rem;
        }

        .logo-section h1 {
            font-size: 2.2rem;
            font-weight: 800;
            background: linear-gradient(135deg, #FFF, var(--text-muted));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            letter-spacing: -1px;
        }

        .logo-section p {
            color: var(--text-muted);
            font-size: 0.95rem;
            margin-top: 0.2rem;
        }

        .stats {
            display: flex;
            gap: 1.5rem;
        }

        .stat-badge {
            background: var(--card-bg);
            border: 1px solid var(--card-border);
            padding: 0.6rem 1.2rem;
            border-radius: 50px;
            font-size: 0.9rem;
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }

        .stat-badge span {
            font-weight: 600;
            color: #FFF;
        }

        /* Filter Tabs */
        .filters {
            max-width: 1200px;
            margin: 0 auto 2rem auto;
            display: flex;
            gap: 0.75rem;
            flex-wrap: wrap;
        }

        .filter-btn {
            background: rgba(255, 255, 255, 0.03);
            border: 1px solid rgba(27, 77, 62, 0.2);
            color: var(--text-muted);
            padding: 0.6rem 1.5rem;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            font-size: 0.9rem;
            transition: var(--transition);
        }

        .filter-btn:hover {
            background: rgba(27, 77, 62, 0.15);
            color: var(--text-main);
            border-color: var(--primary-light);
        }

        .filter-btn.active {
            background: var(--primary);
            color: #FFF;
            border-color: var(--primary);
            box-shadow: 0 4px 15px rgba(27, 77, 62, 0.3);
        }

        /* Grid */
        .grid {
            max-width: 1200px;
            margin: 0 auto;
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(360px, 1fr));
            gap: 1.5rem;
        }

        .card {
            background: var(--card-bg);
            border: 1px solid var(--card-border);
            border-radius: 12px;
            padding: 1.5rem;
            backdrop-filter: blur(10px);
            transition: var(--transition);
            display: flex;
            flex-direction: column;
            justify-content: space-between;
        }

        .card:hover {
            transform: translateY(-4px);
            border-color: var(--primary-light);
            box-shadow: 0 8px 25px rgba(27, 77, 62, 0.15);
        }

        .card-header {
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            margin-bottom: 1rem;
        }

        .company-name {
            font-size: 1.25rem;
            font-weight: 600;
            color: #FFF;
            margin-bottom: 0.25rem;
        }

        .company-meta {
            font-size: 0.85rem;
            color: var(--text-muted);
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }

        .badge-city {
            background: rgba(27, 77, 62, 0.3);
            color: #FFF;
            padding: 0.2rem 0.6rem;
            border-radius: 4px;
            font-size: 0.75rem;
            font-weight: 600;
        }

        .contacted-badge {
            background: #2e7d32;
            color: #fff;
            padding: 0.2rem 0.5rem;
            border-radius: 4px;
            font-size: 0.7rem;
            font-weight: bold;
        }

        .card.contacted {
            opacity: 0.45;
            filter: grayscale(30%);
            border-color: rgba(255, 255, 255, 0.05);
            background: rgba(0, 0, 0, 0.2);
        }

        /* Contact Details */
        .details-section {
            background: rgba(0, 0, 0, 0.2);
            border-radius: 8px;
            padding: 0.8rem;
            margin-bottom: 1.2rem;
            font-size: 0.85rem;
            display: flex;
            flex-direction: column;
            gap: 0.5rem;
        }

        .detail-row {
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .detail-label {
            color: var(--text-muted);
        }

        .detail-value {
            font-weight: 600;
            color: var(--text-main);
            max-width: 200px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }

        /* Actions Section */
        .actions-group {
            display: flex;
            flex-direction: column;
            gap: 0.5rem;
        }

        .action-row {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 0.5rem;
        }

        .btn {
            background: rgba(255, 255, 255, 0.04);
            border: 1px solid rgba(27, 77, 62, 0.2);
            color: var(--text-main);
            padding: 0.6rem;
            border-radius: 6px;
            font-size: 0.8rem;
            font-weight: 600;
            cursor: pointer;
            transition: var(--transition);
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 0.4rem;
            text-decoration: none;
        }

        .btn:hover {
            background: rgba(27, 77, 62, 0.25);
            border-color: var(--primary-light);
        }

        .btn-primary {
            background: var(--primary);
            border-color: var(--primary);
            color: #FFF;
            grid-column: span 2;
        }

        .btn-primary:hover {
            background: var(--primary-light);
            border-color: var(--primary-light);
            box-shadow: 0 4px 10px rgba(27, 77, 62, 0.2);
        }

        /* Toast notifications */
        .toast {
            position: fixed;
            bottom: 2rem;
            right: 2rem;
            background: var(--primary);
            color: #FFF;
            padding: 0.8rem 1.6rem;
            border-radius: 8px;
            font-weight: 600;
            box-shadow: 0 10px 25px rgba(0,0,0,0.5);
            opacity: 0;
            transform: translateY(20px);
            transition: all 0.3s cubic-bezier(0.175, 0.885, 0.32, 1.275);
            pointer-events: none;
            z-index: 9999;
        }

        .toast.show {
            opacity: 1;
            transform: translateY(0);
        }
    </style>
</head>
<body>

    <header>
        <div class="logo-section">
            <h1>Outreach B2B Dashboard</h1>
            <p>Generare de lead-uri calificate și campanii email reci</p>
        </div>
        <div class="stats">
            <div class="stat-badge">Total Firme: <span id="total-count">0</span></div>
            <div class="stat-badge">Cu Email: <span id="email-count">0</span></div>
            <div class="stat-badge">Fără Email: <span id="form-count">0</span></div>
        </div>
    </header>

    <div class="filters">
        <button class="filter-btn active" onclick="filterCity('All')">Toate Orașele</button>
        <button class="filter-btn" onclick="filterCity('New York')">New York</button>
        <button class="filter-btn" onclick="filterCity('London')">London</button>
        <button class="filter-btn" onclick="filterCity('Milano')">Milano</button>
        <button class="filter-btn" onclick="filterCity('București')">București</button>
    </div>

    <div class="grid" id="leads-grid">
        <!-- Cards generated dynamically -->
    </div>

    <div class="toast" id="toast">Copiato in Clipboard!</div>

    <script>
        // Embed the leads data directly
        const leadsData = __LEADS_JSON__;

        function renderLeads(data) {
            const grid = document.getElementById('leads-grid');
            grid.innerHTML = '';
            
            let emailCount = 0;
            let formCount = 0;
            
            const contactedLeads = JSON.parse(localStorage.getItem('contactedLeads') || '{}');
            
            data.forEach((lead, index) => {
                if (lead.email && lead.email !== 'N/A') {
                    emailCount++;
                } else {
                    formCount++;
                }
                
                const isContacted = contactedLeads[lead.company] || false;
                const card = document.createElement('div');
                card.className = `card ${isContacted ? 'contacted' : ''}`;
                card.id = `card-${lead.id}`;
                card.setAttribute('data-city', lead.city_filter);
                
                const hasEmail = lead.email && lead.email !== 'N/A';
                
                const mailtoLink = hasEmail 
                    ? `mailto:${lead.email}?subject=${encodeURIComponent(lead.subject)}&body=${encodeURIComponent(lead.body)}`
                    : '#';

                card.innerHTML = `
                    <div>
                        <div class="card-header">
                            <div>
                                <h3 class="company-name" style="display: flex; align-items: center; gap: 0.5rem; justify-content: space-between;">
                                    ${lead.company}
                                    <span class="contacted-badge" id="badge-${lead.id}" style="display: ${isContacted ? 'inline-block' : 'none'};">Trimis ✓</span>
                                </h3>
                                <div class="company-meta">
                                    <span class="badge-city">${lead.city_filter}</span>
                                    <span>${lead.location}</span>
                                </div>
                            </div>
                        </div>
                        
                        <div class="details-section">
                            <div class="detail-row">
                                <span class="detail-label">Manager/Decident:</span>
                                <span class="detail-value" title="${lead.decident}">${lead.decident}</span>
                            </div>
                            <div class="detail-row">
                                <span class="detail-label">Email:</span>
                                <span class="detail-value" title="${lead.email}">${lead.email}</span>
                            </div>
                            <div class="detail-row">
                                <span class="detail-label">Telefon:</span>
                                <span class="detail-value" title="${lead.phone}">${lead.phone}</span>
                            </div>
                        </div>
                    </div>
                    
                    <div class="actions-group">
                        ${hasEmail ? `
                            <a href="${mailtoLink}" class="btn btn-primary">
                                ✉️ Deschide Email (Zoho / Client Implicit)
                            </a>
                        ` : `
                            <a href="${lead.source_url}" target="_blank" class="btn btn-primary">
                                📝 Deschide Formular Contact Sursă
                            </a>
                        `}
                        <div class="action-row">
                            <button class="btn" onclick="copyLeadField(${lead.id}, 'email', 'Email Copiat!')">
                                📋 Copiază Email
                            </button>
                            <button class="btn" onclick="copyLeadField(${lead.id}, 'subject', 'Subiect Copiat!')">
                                📋 Copiază Subiect
                            </button>
                        </div>
                        
                        <div class="action-row">
                            <button class="btn" style="grid-column: span 2;" onclick="copyLeadField(${lead.id}, 'body', 'Mesaj Copiat!')">
                                📋 Copiază Corp Mesaj Complet
                            </button>
                        </div>
                        
                        <div class="action-row">
                            <a href="${lead.website}" target="_blank" class="btn">🌐 Vizitează Site</a>
                            ${lead.linkedin !== 'N/A' ? `<a href="${lead.linkedin}" target="_blank" class="btn">💼 LinkedIn</a>` : '<button class="btn" disabled style="opacity:0.3">💼 LinkedIn N/A</button>'}
                        </div>
                        
                        <div class="action-row" style="margin-top: 0.5rem;">
                            <button class="btn" id="btn-toggle-${lead.id}" onclick="toggleContacted(${lead.id}, '${cleanJSString(lead.company)}')" style="grid-column: span 2; background: rgba(255, 255, 255, 0.02); border-color: rgba(27, 77, 62, 0.15); color: var(--text-muted); font-size: 0.75rem;">
                                ${isContacted ? '↩️ Resetează Status' : '☑️ Marchează ca Trimis'}
                            </button>
                        </div>
                    </div>
                `;
                grid.appendChild(card);
            });
            
            document.getElementById('total-count').innerText = data.length;
            document.getElementById('email-count').innerText = emailCount;
            document.getElementById('form-count').innerText = formCount;
        }

        function cleanJSString(str) {
            return str.replace(/\\\\/g, '\\\\\\\\').replace(/'/g, "\\\\'").replace(/\\n/g, '\\\\n').replace(/\\r/g, '');
        }

        function filterCity(city) {
            // Update active button styling
            const buttons = document.querySelectorAll('.filter-btn');
            buttons.forEach(btn => {
                if (btn.innerText.includes(city) || (city === 'All' && btn.innerText.includes('Toate'))) {
                    btn.classList.add('active');
                } else {
                    btn.classList.remove('active');
                }
            });
            
            if (city === 'All') {
                renderLeads(leadsData);
            } else {
                const filtered = leadsData.filter(lead => lead.city_filter === city);
                renderLeads(filtered);
            }
        }

        function toggleContacted(id, company) {
            const contactedLeads = JSON.parse(localStorage.getItem('contactedLeads') || '{}');
            const isContacted = !contactedLeads[company];
            
            if (isContacted) {
                contactedLeads[company] = true;
            } else {
                delete contactedLeads[company];
            }
            
            localStorage.setItem('contactedLeads', JSON.stringify(contactedLeads));
            
            const card = document.getElementById(`card-${id}`);
            const badge = document.getElementById(`badge-${id}`);
            const button = document.getElementById(`btn-toggle-${id}`);
            
            if (isContacted) {
                card.classList.add('contacted');
                badge.style.display = 'inline-block';
                button.innerHTML = '↩️ Resetează Status';
                showToast('Marcat ca trimis!');
            } else {
                card.classList.remove('contacted');
                badge.style.display = 'none';
                button.innerHTML = '☑️ Marchează ca Trimis';
                showToast('Status resetat!');
            }
        }

        function copyLeadField(id, field, message) {
            const lead = leadsData.find(l => l.id === id);
            if (lead) {
                copyText(lead[field], message);
            }
        }

        function copyText(text, message) {
            if (text === 'N/A') {
                showToast('Nu există date de copiat!');
                return;
            }
            navigator.clipboard.writeText(text).then(() => {
                showToast(message);
            }).catch(err => {
                console.error('Eroare la copiere: ', err);
            });
        }

        function showToast(message) {
            const toast = document.getElementById('toast');
            toast.innerText = message;
            toast.classList.add('show');
            setTimeout(() => {
                toast.classList.remove('show');
            }, 2500);
        }

        // Initial render
        renderLeads(leadsData);
    </script>
</body>
</html>
"""

    # Escape leads data and insert into html
    leads_json = []
    for idx, lead in enumerate(leads):
        clean_name = clean_decident_name(lead["decident"])
        # Update subject and body using cleaned name
        subject, body = get_email_templates(lead["company"], clean_name, lead["city_filter"].lower())
        leads_json.append({
            "id": idx,
            "company": lead["company"],
            "location": lead["location"],
            "city_filter": lead["city_filter"],
            "website": lead["website"],
            "email": lead["email"],
            "phone": lead["phone"],
            "decident": clean_name,
            "linkedin": lead["linkedin"],
            "source_url": lead["source_url"],
            "subject": subject,
            "body": body
        })
        
    leads_json_str = json.dumps(leads_json)
    html_content = html_template.replace("__LEADS_JSON__", leads_json_str)
    
    # Save outreach dashboard to Desktop
    dashboard_path = os.path.join(desktop_dir, "OUTREACH_DASHBOARD.html")
    with open(dashboard_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
        
    print(f"\n=======================================================")
    print(f"[+] PANOU DE CONTROL GENERAT PE DESKTOP!")
    print(f"   Fisier: Desktop/OUTREACH_DASHBOARD.html")
    print(f"   - Deschide-l in browser (Dublu-click)")
    print(f"   - Copiaza campurile cu 1-click")
    print(f"=======================================================\n")

if __name__ == "__main__":
    main()
