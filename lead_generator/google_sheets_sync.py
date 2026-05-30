import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re

class LeadExportSyncManager:
    """
    Manages lead exports. Supports direct Google Sheets synchronization
    via gspread, and falls back to a premium local Excel file with
    an "Emerald Green" theme if API credentials are not set up.
    """

    def __init__(self, credentials_filename="credentials.json"):
        self.credentials_path = credentials_filename

    def sync_to_google_sheets(self, data, sheet_name_or_url):
        """
        Attempts to authenticate with Google Drive/Sheets API using gspread
        and exports the lead data.
        """
        if not os.path.exists(self.credentials_path):
            raise FileNotFoundError("Google credentials.json file not found.")

        # Import gspread only when needed to prevent dependency errors if not installed
        import gspread
        
        # Authenticate using the service account key
        gc = gspread.service_account(filename=self.credentials_path)
        
        # Open sheet by URL or by title
        if sheet_name_or_url.startswith("https://"):
            sh = gc.open_by_url(sheet_name_or_url)
        else:
            sh = gc.open(sheet_name_or_url)
            
        ws = sh.get_worksheet(0) # Get the first worksheet
        
        # Prepare rows
        headers = ["Business Name", "Phone", "Address", "Rating (1-5)", "Reviews Count", "Website", "Profile URL"]
        rows = [headers]
        for item in data:
            rows.append([
                item.get("Business Name", ""),
                item.get("Phone", ""),
                item.get("Address", ""),
                item.get("Rating (1-5)", 0.0),
                item.get("Reviews Count", 0),
                item.get("Website", ""),
                item.get("Profile URL", "")
            ])
            
        # Clear sheet and write all rows in one batch call
        ws.clear()
        ws.update("A1", rows)
        return sh.url

    def export_to_local_excel(self, data, output_path):
        """
        Generates a premium, "Emerald Green" themed Excel sheet as a robust local backup.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "B2B Sales Leads"
        ws.views.sheetView[0].showGridLines = True
        
        headers = ["Business Name", "Phone", "Address", "Rating (1-5)", "Reviews Count", "Website Link", "Profile Link"]
        ws.append(headers)
        
        for item in data:
            website = item.get("Website", "")
            profile_url = item.get("Profile URL", "")
            
            ws.append([
                item.get("Business Name", ""),
                item.get("Phone", ""),
                item.get("Address", ""),
                item.get("Rating (1-5)", 0.0),
                item.get("Reviews Count", 0),
                "Visit Site ↗" if website else "N/A",
                "Yellowpages Profile ↗" if profile_url else "N/A"
            ])
            
            # Set native Excel hyperlink properties for robust compatibility!
            current_row = ws.max_row
            if website:
                ws.cell(row=current_row, column=6).hyperlink = website
            if profile_url:
                ws.cell(row=current_row, column=7).hyperlink = profile_url

        # Emerald Green Palette Styling
        header_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid") # Dark Emerald
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        zebra_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid") # Soft Emerald Tint
        normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border_side = Side(border_style="thin", color="A7F3D0")
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        num_rows = len(data)
        data_start_row = 2
        data_end_row = num_rows + 1

        # Header Row Styling (Height 28)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center if col_idx != 1 and col_idx != 3 else align_left
            cell.border = cell_border

        # Data Rows Styling (Height 32 for spacious airy spacing)
        for r_idx in range(data_start_row, data_end_row + 1):
            ws.row_dimensions[r_idx].height = 32
            row_fill = zebra_fill if r_idx % 2 == 0 else normal_fill
            
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.fill = row_fill
                cell.border = cell_border
                
                if c_idx == 1: # Business Name
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = align_left
                elif c_idx == 2: # Phone
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = align_center
                elif c_idx == 3: # Address
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = align_left
                elif c_idx == 4: # Rating
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = align_center
                    cell.number_format = '0.0'
                elif c_idx == 5: # Reviews
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = align_center
                    cell.number_format = '0'
                else: # Links (Columns 6 and 7)
                    cell.alignment = align_center
                    if cell.value == "N/A":
                        cell.font = Font(name="Calibri", size=10, italic=True, color="A0AEC0")
                    else:
                        cell.font = Font(name="Calibri", size=10, underline="single", color="065F46")


        # Summary Row (Formulas)
        summary_row = data_end_row + 1
        ws.row_dimensions[summary_row].height = 24
        
        summary_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        summary_font = Font(name="Calibri", size=10, bold=True, color="065F46")
        double_bottom_border = Border(
            top=Side(border_style="thin", color="34D399"),
            bottom=Side(border_style="double", color="065F46"), # Accounting Double Underline
            left=thin_border_side,
            right=thin_border_side
        )

        ws.cell(row=summary_row, column=1, value="Average / Totals").alignment = align_left
        ws.cell(row=summary_row, column=2, value="").alignment = align_center
        ws.cell(row=summary_row, column=3, value="").alignment = align_left
        ws.cell(row=summary_row, column=4, value=f"=AVERAGE(D2:D{data_end_row})").number_format = '0.0'
        ws.cell(row=summary_row, column=5, value=f"=SUM(E2:E{data_end_row})").number_format = '#,##0'
        ws.cell(row=summary_row, column=6, value="Total Leads").alignment = align_right
        ws.cell(row=summary_row, column=7, value=f"=COUNTA(A{data_start_row}:A{data_end_row})").number_format = '0'

        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=summary_row, column=c_idx)
            cell.fill = summary_fill
            cell.font = summary_font
            cell.border = double_bottom_border
            if c_idx in [4, 5, 7]:
                cell.alignment = align_center if c_idx in [4, 5] else align_right

        # Auto-Fit Columns (Calculated based on actual display text)
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith('='):
                    val_str = "Average / Totals"
                max_len = max(max_len, len(val_str))
                
            if col[0].column in [1, 3]: # Name or Address
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 18), 45)
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)



        wb.save(output_path)
        print(f"Emerald Green Excel backup saved successfully to: {output_path}")

# Self-test block
if __name__ == "__main__":
    print("Testing Lead Sync/Export Manager...")
    manager = LeadExportSyncManager()
    test_data = [
        {"Business Name": "Beacon Hill Dental", "Phone": "(617) 555-0199", "Address": "12 Beacon St, Boston, MA 02108", "Rating (1-5)": 4.5, "Reviews Count": 28, "Website": "http://beaconhilldental.com", "Profile URL": "http://yp.com/1"},
        {"Business Name": "Boston Dental Group", "Phone": "(617) 555-0144", "Address": "100 Boylston St, Boston, MA 02116", "Rating (1-5)": 5.0, "Reviews Count": 142, "Website": "http://bostondental.com", "Profile URL": "http://yp.com/2"}
    ]
    test_file = "test_leads.xlsx"
    manager.export_to_local_excel(test_data, test_file)
    if os.path.exists(test_file):
        os.remove(test_file)
        print("Self-test completed successfully!")
