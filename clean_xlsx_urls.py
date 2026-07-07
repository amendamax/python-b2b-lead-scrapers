import os
import glob
import openpyxl
import urllib.parse

def clean_url(url_str):
    if not url_str:
        return url_str
    url_str = str(url_str).strip()
    if not url_str.startswith('http'):
        return url_str
    
    # Parse URL and strip query/fragment
    try:
        parsed = urllib.parse.urlparse(url_str)
        # Reconstruct base URL without query parameters
        clean = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
        # Make sure it ends nicely
        return clean
    except Exception:
        # Fallback to simple split on '?'
        return url_str.split('?')[0]

def main():
    workspace_dir = r"C:\Users\bratu\Documents\antigravity\amazing-borg"
    
    excel_files = glob.glob(os.path.join(workspace_dir, "leads_*.xlsx"))
    outreach_file = os.path.join(workspace_dir, "PROSPECTE_AGENTII_OUTREACH.xlsx")
    if os.path.exists(outreach_file):
        excel_files.append(outreach_file)
        
    print(f"Cleaning URLs in {len(excel_files)} Excel spreadsheets...")
    
    for file_path in excel_files:
        if "contacted_success" in os.path.basename(file_path).lower():
            continue
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Find website column
            web_col = None
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for c_idx, h in enumerate(headers, start=1):
                h_str = str(h).lower()
                if "site" in h_str or "website" in h_str:
                    web_col = c_idx
                    break
                    
            if not web_col:
                # Fallback: check second row
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=2, column=c).value)
                    if val.startswith('http'):
                        web_col = c
                        break
                        
            if web_col:
                cleaned_count = 0
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=web_col)
                    original_val = cell.value
                    original_hyperlink = cell.hyperlink.target if cell.hyperlink else None
                    
                    target_url = original_hyperlink or original_val
                    if target_url and "?" in str(target_url):
                        cleaned_val = clean_url(target_url)
                        
                        # Update cell value/hyperlink
                        if cell.hyperlink:
                            cell.value = "Vizitează Site ↗"
                            cell.hyperlink = cleaned_val
                        else:
                            cell.value = cleaned_val
                        cleaned_count += 1
                        
                if cleaned_count > 0:
                    wb.save(file_path)
                    print(f"  {os.path.basename(file_path)}: Cleaned {cleaned_count} URLs.")
            else:
                print(f"  {os.path.basename(file_path)}: No website column found.")
                
        except Exception as e:
            print(f"  Error processing {os.path.basename(file_path)}: {e}")
            
    print("\nURL cleaning completed!")

if __name__ == "__main__":
    main()
