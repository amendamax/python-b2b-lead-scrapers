import os
import glob
import re
import openpyxl
from curl_cffi import requests
from bs4 import BeautifulSoup
import concurrent.futures
import urllib.parse
import sys

# Reconfigure stdout for UTF-8
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except AttributeError:
    pass

def clean_domain(url):
    try:
        parsed = urllib.parse.urlparse(url)
        return parsed.netloc.replace("www.", "").lower()
    except:
        return ""

def find_contact_subpages(homepage_url, html_content):
    subpages = []
    keywords = ["contact", "about", "team", "contatt", "despre", "noi", "info", "legal", "privacy"]
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        for a in soup.find_all('a', href=True):
            href = a['href'].strip()
            text = a.get_text().strip().lower()
            href_low = href.lower()
            
            is_match = any(kw in href_low or kw in text for kw in keywords)
            if is_match:
                if href_low.startswith('mailto:') or href_low.startswith('tel:') or href_low.startswith('javascript:'):
                    continue
                if any(x in href_low for x in ["facebook.com", "linkedin.com", "twitter.com", "instagram.com"]):
                    continue
                
                absolute_url = urllib.parse.urljoin(homepage_url, href)
                # Keep it within same domain
                if clean_domain(absolute_url) == clean_domain(homepage_url):
                    subpages.append(absolute_url)
    except:
        pass
    return list(set(subpages))

def check_site_for_email(url, target_email):
    """Crawls homepage and up to 2 subpages to verify if the target email is listed."""
    session = requests.Session(impersonate="chrome110")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    target_email = target_email.strip().lower()
    try:
        r = session.get(url, headers=headers, timeout=10)
        html = r.text
        
        # 1. Search in homepage text/HTML
        emails = [e.lower() for e in re.findall(r'[\w\.-]+@[\w\.-]+\.\w+', html)]
        if target_email in emails:
            return True
            
        # 2. Try subpages if not found on homepage
        subpages = find_contact_subpages(url, html)
        for sub in subpages[:2]:
            try:
                sub_r = session.get(sub, headers=headers, timeout=8)
                sub_emails = [e.lower() for e in re.findall(r'[\w\.-]+@[\w\.-]+\.\w+', sub_r.text)]
                if target_email in sub_emails:
                    return True
            except:
                pass
    except Exception as e:
        # If site fails completely, we try HTTP fallback
        if url.startswith("https://"):
            url_http = url.replace("https://", "http://")
            try:
                r = session.get(url_http, headers=headers, timeout=8)
                emails = [e.lower() for e in re.findall(r'[\w\.-]+@[\w\.-]+\.\w+', r.text)]
                if target_email in emails:
                    return True
            except:
                pass
                
    return False

def process_lead(lead_info):
    company, url, email, file_path, row_idx = lead_info
    print(f"Verifying {company} ({email})...")
    is_real = check_site_for_email(url, email)
    if is_real:
        print(f"  [VERIFIED] {company} email is REAL.")
    else:
        print(f"  [FAKE/GUESSED] {company} email was NOT found on their site.")
    return row_idx, email, is_real, file_path, company

def main():
    workspace_dir = r"C:\Users\bratu\Documents\antigravity\amazing-borg"
    excel_files = glob.glob(os.path.join(workspace_dir, "leads_*.xlsx"))
    outreach_file = os.path.join(workspace_dir, "PROSPECTE_AGENTII_OUTREACH.xlsx")
    if os.path.exists(outreach_file):
        excel_files.append(outreach_file)
        
    leads_to_verify = []
    ignore_domains = ['example.com', 'domain.com', 'yourcompany.co.uk', 'yourdomain.com', 'yourdomain', 'example']
    
    print("Gathering leads to verify from Excel files...")
    for file_path in excel_files:
        if "contacted_success" in os.path.basename(file_path).lower():
            continue
            
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Find column indices
            email_col = None
            web_col = None
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for c_idx, h in enumerate(headers, start=1):
                h_str = str(h).lower()
                if "email" in h_str or "e-mail" in h_str:
                    email_col = c_idx
                elif "site" in h_str or "website" in h_str:
                    web_col = c_idx
                    
            if not email_col or not web_col:
                # Fallback search
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=2, column=c).value)
                    if "@" in val and not email_col:
                        email_col = c
                    elif val.startswith("http") and not web_col:
                        web_col = c
                        
            if email_col and web_col:
                for row in range(2, ws.max_row + 1):
                    company = ws.cell(row=row, column=1).value
                    if not company:
                        continue
                    cell_web = ws.cell(row=row, column=web_col)
                    website = cell_web.hyperlink.target if cell_web.hyperlink else cell_web.value
                    email = ws.cell(row=row, column=email_col).value
                    
                    if email and website and "@" in str(email):
                        email_str = str(email).strip()
                        is_placeholder = any(ig in email_str.lower() for ig in ignore_domains)
                        if not is_placeholder:
                            leads_to_verify.append((company, website, email_str, file_path, row))
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")
            
    print(f"\nFound {len(leads_to_verify)} leads to verify. Starting verification with 10 threads...")
    
    verified_leads = {} # file_path -> list of (row_idx, is_real)
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        results = executor.map(process_lead, leads_to_verify)
        
        for row_idx, email, is_real, file_path, company in results:
            if file_path not in verified_leads:
                verified_leads[file_path] = []
            verified_leads[file_path].append((row_idx, is_real, email, company))
            
    # Now update Excel files: if email is not real, change cell value to N/A or delete it!
    print("\nUpdating Excel spreadsheets with verification results...")
    for file_path, updates in verified_leads.items():
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Find email column index
            email_col = None
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for c_idx, h in enumerate(headers, start=1):
                h_str = str(h).lower()
                if "email" in h_str or "e-mail" in h_str:
                    email_col = c_idx
                    break
            if not email_col:
                for c in range(1, ws.max_column + 1):
                    if "@" in str(ws.cell(row=2, column=c).value):
                        email_col = c
                        break
                        
            if email_col:
                saved = False
                for row_idx, is_real, email, company in updates:
                    if not is_real:
                        # Email was fake/guessed. Delete it!
                        ws.cell(row=row_idx, column=email_col, value="N/A")
                        saved = True
                        print(f"  {os.path.basename(file_path)}: Cleared fake email for {company}.")
                if saved:
                    wb.save(file_path)
        except Exception as e:
            print(f"Error updating file {file_path}: {e}")
            
    print("\nVerification and database cleanup completed!")

if __name__ == "__main__":
    main()
